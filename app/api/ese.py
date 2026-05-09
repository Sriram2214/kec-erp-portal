from flask import jsonify, request, send_file
from flask_login import login_required, current_user
from app.api import api
from app.models import (Student, Course, ExamSchedule, Attendance, CourseRegistration, Curriculum,
                        AcademicYear, DummySticker, FoilMark, FeeClearance, Department, 
                        Degree, Batch, Regulation)
from app import db, limiter
from app.utils.logger import audit_log
import datetime as dt
import os
import io
import math
import logging
import random
import string
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm, inch
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                Paragraph, Spacer, HRFlowable, PageBreak, KeepTogether, Image as RLImage)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from reportlab.lib import colors

# ── Constants & Helpers ──────────────────────────────────────────────
NAVY       = colors.HexColor('#1a2a5e')
GOLD       = colors.HexColor('#c9a227')
RED        = colors.HexColor('#dc2626')
WHITE      = colors.white
BLACK      = colors.black

# Resolve paths relative to this file's location (app/api/) → go up 2 dirs to project root
from app.utils.pdf import get_institutional_header, HEADER_IMG, LOGO_PATH



# ── ESE Attendance ──


@api.route('/ese/courses-by-date', methods=['GET'])
@login_required
def get_courses_by_date():
    date_str = request.args.get('date', '')
    if not date_str:
        return jsonify({'message': 'Date required'}), 400
    try:
        target_date = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        return jsonify({'message': 'Invalid date format (YYYY-MM-DD)'}), 400
    
    from sqlalchemy.orm import joinedload
    schedules = ExamSchedule.query.options(joinedload(ExamSchedule.course).joinedload(Course.curriculum).joinedload(Curriculum.department)).filter_by(exam_date=target_date).all()
    results = []
    for s in schedules:
        sticker_count = DummySticker.query.filter_by(exam_schedule_id=s.id).count()
        results.append({
            'schedule_id': s.id,
            'course_id': s.course.id,
            'course_code': s.course.course_code,
            'course_title': s.course.course_title,
            'session': s.session,
            'venue': s.venue,
            'department': s.course.curriculum.department.code,
            'sticker_count': sticker_count
        })
    return jsonify(results)

# ── GET /api/ese/students?course_code=XXX
# ── GET /api/exam-attendance/<course_code>  (alias)
@api.route('/ese/students', methods=['GET'])
@api.route('/exam-attendance/<path:dummy>', methods=['GET'])
@login_required
def ese_students(dummy=None):
    course_id = request.args.get('course_id')
    if not course_id:
        return jsonify({'message': 'course_id is MANDATORY'}), 400
    
    course = Course.query.get_or_404(course_id)
    schedule = ExamSchedule.query.filter_by(course_id=course.id).first()

    print("=" * 80)
    print("[DEBUG] PDF/API ROUTE STARTED: ese_students")
    print(f"[DEBUG] course_id param = {course_id}")
    print(f"[DEBUG] course_code param = {request.args.get('course_code')}")
    print(f"[DEBUG] Resolved Course ID = {course.id}")
    print(f"[DEBUG] Resolved Course Code = {course.course_code}")
    print(f"[DEBUG] Curriculum ID = {course.curriculum_id}")
    print(f"[DEBUG] Department = {course.curriculum.department.code}")
    print(f"[DEBUG] Regulation = {course.curriculum.regulation.name}")
    print(f"[DEBUG] Batch = {course.curriculum.batch.label}")
    # ── Optimized Load existing attendance and stickers ──
    existing = {}
    stickers = {}
    if schedule:
        # Batch fetch for speed
        att_data = db.session.query(Attendance.student_id, Attendance.status).filter(Attendance.exam_schedule_id == schedule.id).all()
        existing = {a[0]: a[1] for a in att_data}
        
        ds_data = db.session.query(DummySticker.student_id, DummySticker.dummy_number).filter(DummySticker.dummy_number.like(f"{dt.datetime.now().year % 100}%")).all()
        # Note: We need stickers for THIS schedule specifically
        ds_schedule = db.session.query(DummySticker.student_id, DummySticker.dummy_number).filter(DummySticker.exam_schedule_id == schedule.id).all()
        stickers = {d[0]: d[1] for d in ds_schedule}

    # ── STRICT CourseRegistration Query ONLY ──
    students = (
        db.session.query(Student)
        .join(
            CourseRegistration,
            CourseRegistration.student_id == Student.id
        )
        .filter(
            CourseRegistration.course_id == course.id
        )
        .order_by(Student.register_number)
        .all()
    )

    print(f"[FINAL STRICT MODE] Course ID: {course.id} | Students: {len(students)}")

    # ── High-Performance Dummy Sticker Generation (Only if missing) ──
    # [PERFORMANCE] If on Vercel, we might want to skip auto-gen if student count is high to avoid 10s timeout
    if schedule and len(students) < 100: 
        to_generate = [s for s in students if s.id not in stickers]
        if to_generate:
            print(f"[DUMMY GEN] Auto-generating {len(to_generate)} stickers...")
            year_prefix = str(dt.datetime.now().year % 100)
            used_dummy_nos = {d[0] for d in db.session.query(DummySticker.dummy_number).filter(DummySticker.dummy_number.like(f"{year_prefix}%")).all()}
            
            new_stickers = []
            for s in to_generate:
                dept_prefix = "".join(filter(str.isalnum, (s.department or 'GEN')))[:3].upper()
                while True:
                    rand_part = ''.join(random.choices(string.digits, k=5))
                    dummy_no = f"{year_prefix}{dept_prefix}{rand_part}"
                    if dummy_no not in used_dummy_nos:
                        used_dummy_nos.add(dummy_no)
                        break
                
                new_stickers.append(DummySticker(
                    student_id=s.id,
                    exam_schedule_id=schedule.id,
                    dummy_number=dummy_no,
                    foil_number=str(random.randint(10000, 99999))
                ))
                stickers[s.id] = dummy_no 
            
            if new_stickers:
                try:
                    db.session.bulk_save_objects(new_stickers)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    logging.error(f"DUMMY GEN ERROR: {e}")

    return jsonify({
        'course_id':    course.id,
        'course_code':  course.course_code,
        'course_title': course.course_title,
        'department':   course.curriculum.department.code if course.curriculum and course.curriculum.department else 'N/A',
        'semester':     course.semester,
        'exam_date':    schedule.exam_date.strftime('%d.%m.%Y') if schedule and schedule.exam_date else None,
        'session':      schedule.session if schedule else None,
        'venue':        schedule.venue if schedule else None,
        'schedule_id':  schedule.id if schedule else None,
        'total':        len(students),
        'students': [{
            'id':              s.id,
            'register_number': s.register_number,
            'name':            s.name,
            'department':      s.department,
            'batch':           s.batch,
            'semester':        s.semester,
            'status':          existing.get(s.id, 'Present'),
            'dummy_number':    stickers.get(s.id, '—')
        } for s in students]
    })

@api.route('/ese/auto-enroll', methods=['POST'])
@login_required
def ese_auto_enroll():
    if current_user.role not in ['admin', 'coe']:
        return jsonify({'message': 'Access denied'}), 403
    
    data = request.get_json()
    course_id = data.get('course_id')
    if not course_id:
        return jsonify({'message': 'course_id required'}), 400
    
    course = Course.query.get_or_404(course_id)
    curr = course.curriculum
    
    # Find all students matching this curriculum (Batch + Dept + Regulation)
    students = Student.query.filter(
        Student.batch == curr.batch.label,
        Student.regulation == curr.regulation.name,
        Student.department == curr.department.code
    ).all()
    
    if not students:
        return jsonify({'message': 'No students found matching this curriculum (Batch/Dept mismatch)'}), 404
        
    # Check existing registrations
    existing_sids = {r.student_id for r in CourseRegistration.query.filter_by(course_id=course.id).all()}
    
    to_add = []
    for s in students:
        if s.id not in existing_sids:
            to_add.append(CourseRegistration(student_id=s.id, course_id=course.id))
            
    if to_add:
        db.session.bulk_save_objects(to_add)
        db.session.commit()
        audit_log.log('AUTO_ENROLL', {'course': course.course_code, 'count': len(to_add)})
        
    return jsonify({
        'message': f'Successfully enrolled {len(to_add)} students.',
        'enrolled': len(to_add)
    })


@api.route('/ese/attendance', methods=['POST'])
@api.route('/save-attendance', methods=['POST'])
@login_required
def save_ese_attendance():
    data = request.get_json()
    course_id = data.get('course_id')
    entries   = data.get('entries', [])

    if not course_id:
        return jsonify({'message': 'course_id required'}), 400
    if not entries:
        return jsonify({'message': 'No attendance entries provided'}), 400

    course = Course.query.get_or_404(course_id)

    # ── Find or auto-create exam schedule ────────────────────────────
    schedule = ExamSchedule.query.filter_by(course_id=course.id).first()
    if not schedule:
        exam_date_str = data.get('exam_date', '')
        session_val   = data.get('session', 'FN')
        try:
            edate = dt.datetime.strptime(exam_date_str, '%Y-%m-%d').date() if exam_date_str else dt.date.today()
        except Exception:
            edate = dt.date.today()
        ay = AcademicYear.query.filter_by(is_current=True).first()
        schedule = ExamSchedule(
            course_id        = course.id,
            exam_date        = edate,
            session          = session_val,
            academic_year_id = ay.id if ay else None,
        )
        db.session.add(schedule)
        db.session.flush()

    # ── High-Performance Bulk Upsert ──────────────────────────────────
    # 1. Fetch all existing records for this schedule in ONE query
    existing_records = Attendance.query.filter_by(exam_schedule_id=schedule.id).all()
    existing_map = {att.student_id: att for att in existing_records}
    
    saved = 0
    to_add = []
    
    for entry in entries:
        sid    = entry.get('student_id')
        status = entry.get('status', 'Present')
        if not sid: continue

        if sid in existing_map:
            # Update existing in-memory object
            existing_map[sid].status = status
        else:
            # Prepare new object for bulk insertion
            to_add.append(Attendance(student_id=sid, exam_schedule_id=schedule.id, status=status))
        saved += 1

    try:
        if to_add:
            db.session.bulk_save_objects(to_add)
        
        db.session.commit()
        audit_log.log('SAVE_ESE_ATTENDANCE', {'course': course.course_code, 'count': saved, 'by': current_user.username})
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Save failed: {str(e)}'}), 500

    absent_count = sum(1 for e in entries if e.get('status') == 'Absent')
    mp_count     = sum(1 for e in entries if e.get('status') == 'Malpractice')
    return jsonify({
        'message': f'{saved} attendance record(s) saved.',
        'saved': saved,
        'absent': absent_count,
        'malpractice': mp_count,
        'schedule_id': schedule.id
    })

@api.route('/ese/attendance-pdf', methods=['GET'])
@login_required
def ese_attendance_pdf():
    course_id = request.args.get('course_id')
    if not course_id:
        return jsonify({'message': 'course_id is MANDATORY for PDF generation'}), 400
    
    course = Course.query.get_or_404(course_id)

    print("=" * 80)
    print("[DEBUG] PDF/API ROUTE STARTED: ese_attendance_pdf")
    print(f"[DEBUG] course_id param = {course_id}")
    print(f"[DEBUG] course_code param = {request.args.get('course_code')}")
    print(f"[DEBUG] Resolved Course ID = {course.id}")
    print(f"[DEBUG] Resolved Course Code = {course.course_code}")
    print(f"[DEBUG] Curriculum ID = {course.curriculum_id}")
    print(f"[DEBUG] Department = {course.curriculum.department.code}")
    print(f"[DEBUG] Regulation = {course.curriculum.regulation.name}")
    print(f"[DEBUG] Batch = {course.curriculum.batch.label}")
    print("=" * 80)

    schedule = ExamSchedule.query.filter_by(course_id=course.id).first()
    ay = AcademicYear.query.filter_by(is_current=True).first()
    existing = {}
    if schedule:
        for att in Attendance.query.filter_by(exam_schedule_id=schedule.id).all():
            existing[att.student_id] = att.status

    # ── STRICT CourseRegistration Query ──
    students = (
        db.session.query(Student)
        .join(CourseRegistration, CourseRegistration.student_id == Student.id)
        .filter(CourseRegistration.course_id == course.id)
        .order_by(Student.register_number)
        .all()
    )
    
    print(f"[DEBUG] Students Loaded = {len(students)}")
    for s in students[:20]:
        print(
            f"[DEBUG STUDENT] "
            f"{s.register_number} | "
            f"{s.department} | "
            f"{s.batch} | "
            f"{s.regulation}"
        )
    print("=" * 80)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=10*mm, bottomMargin=10*mm,
    )

    NAVY  = colors.HexColor('#1a2a5e')
    GOLD  = colors.HexColor('#c9a227')
    WHITE = colors.white
    RED   = colors.HexColor('#dc2626')
    GREY  = colors.HexColor('#64748b')

    styles = getSampleStyleSheet()
    def sty(name, **kw):
        return ParagraphStyle(name, **kw)

    sheet_title= sty('ST', fontName='Helvetica-Bold', fontSize=10, textColor=colors.black, alignment=TA_CENTER, spaceAfter=4)
    label_sty  = sty('LB', fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.black)
    val_sty    = sty('VL', fontName='Helvetica-Bold', fontSize=8.5, textColor=RED) # Matching red course code in screenshot
    val_black  = sty('VB', fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.black)
    
    # White box styling for booklet number entries (mimicking the grid boxes)
    box_cell_style = [('GRID', (0,0), (-1,-1), 0.5, colors.black), ('BACKGROUND', (0,0), (-1,-1), WHITE)]

    exam_date_disp = schedule.exam_date.strftime('%d.%m.%Y %a').upper() if schedule else ''
    session_disp   = schedule.session if schedule else ''

    page_w = A4[0] - 20*mm
    story = []

    get_institutional_header(story, page_w)

    ay_label = ay.label if ay else '2025-26'
    sem_label = ay.semester if ay else 'EVEN'
    # Mocking "APRIL/MAY - 2026" based on current year or semester
    exam_period = "APRIL/MAY - 2026" if sem_label.lower() == 'even' else "NOV/DEC - 2025"
    
    story.append(Paragraph(f'END SEMESTER THEORY EXAMINATIONS – {exam_period}', sheet_title))
    story.append(Paragraph('ATTENDANCE SHEET', sty('ATT', fontName='Helvetica-Bold', fontSize=11, textColor=colors.black, alignment=TA_CENTER, spaceAfter=4)))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.black))
    story.append(Spacer(1, 2*mm))

    dept_name = course.curriculum.department.code
    info_data = [
        [
            Paragraph('<b>Degree &amp; Branch</b> :', label_sty),
            Paragraph(f'<font color="#FFFFFF" backColor="#b91c1c"><b>&nbsp; {dept_name} &nbsp;</b></font>', val_black), # White text on Red bg for dept highlight
            Paragraph('<b>Course Code :</b>', label_sty),
            Paragraph(f'<b>{course.course_code}</b>', val_sty),
        ],
        [
            Paragraph('<b>Exam Date &amp; Session :</b>', label_sty),
            Paragraph(f'<b>{exam_date_disp} – {session_disp}</b>' if exam_date_disp else '<i>Not Scheduled</i>', val_black),
            Paragraph('<b>Course Title :</b>', label_sty),
            Paragraph(f'<b>{course.course_title}</b>', val_black),
        ],
    ]
    info_tbl = Table(info_data, colWidths=[page_w*0.18, page_w*0.37, page_w*0.15, page_w*0.30])
    info_tbl.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 3)
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 2*mm))

    # Column widths adjusted for Answer Booklet No. columns (likely 10 or 12 small boxes)
    col_widths = [page_w*0.05, page_w*0.14, page_w*0.25, page_w*0.25, page_w*0.13, page_w*0.18]
    
    header_row = [
        Paragraph('<b>SNo</b>', sty('TH', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER)),
        Paragraph('<b>Register No</b>', sty('TH2', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER)),
        Paragraph('<b>Name of the Student</b>', sty('TH3', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER)),
        Paragraph('<b>Answer Booklet No.</b>', sty('TH4', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER)),
        Paragraph('<b>"AB" for Absent</b>', sty('TH5', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER)),
        Paragraph('<b>Student Signature</b>', sty('TH6', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER)),
    ]

    # ── High-Performance Table Build ──
    table_data = [header_row]
    booklet_col_w = (page_w * 0.25) / 10.5
    
    for i, s in enumerate(students, 1):
        status = existing.get(s.id, 'Present')
        ab_mark = ''
        if status == 'Absent':
            ab_mark = Paragraph(
                '<b>AB</b>',
                sty(f'AB{i}', fontName='Helvetica-Bold', fontSize=9, textColor=RED, alignment=TA_CENTER)
            )
        elif status == 'Malpractice':
            ab_mark = Paragraph(
                '<b>MP</b>',
                sty(f'MP{i}', fontName='Helvetica-Bold', fontSize=9, textColor=RED, alignment=TA_CENTER)
            )
        
        # CREATE BRAND NEW BOX TABLE EVERY ROW to avoid ReportLab memory reuse artifacts
        booklet_boxes = Table(
            [[''] * 10],
            colWidths=[booklet_col_w] * 10
        )
        booklet_boxes.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('BOTTOMPADDING', (0,0), (-1,-1), 1),
            ('TOPPADDING', (0,0), (-1,-1), 1),
            ('LEFTPADDING', (0,0), (-1,-1), 1),
            ('RIGHTPADDING', (0,0), (-1,-1), 1),
        ]))

        table_data.append([
            Paragraph(str(i), sty(f'SN{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
            Paragraph(s.register_number, sty(f'RN{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
            Paragraph(s.name, sty(f'NM{i}', fontName='Helvetica', fontSize=8)),
            booklet_boxes,
            ab_mark,
            '',
        ])

    print(f"[FINAL PDF ROW COUNT] {len(table_data)-1}")

    att_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    att_table.setStyle(TableStyle([
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 8),
        ('ALIGN',      (0,0), (-1,0), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.black),
        ('BOTTOMPADDING', (0,1), (-1,-1), 2),
        ('TOPPADDING', (0,1), (-1,-1), 2),
    ]))
    story.append(att_table)
    story.append(Spacer(1, 5*mm))

    doc.build(story)
    buf.seek(0)
    filename = f'ESE_Attendance_{course.course_code}_{course.curriculum.department.code}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')

@api.route('/ese/cover-sheet-pdf', methods=['GET'])
@login_required
def ese_cover_sheet_pdf():
    course_id = request.args.get('course_id')
    if not course_id:
        return jsonify({'message': 'course_id is MANDATORY for PDF generation'}), 400
        
    course = Course.query.get_or_404(course_id)

    print("=" * 80)
    print("[DEBUG] PDF/API ROUTE STARTED: ese_cover_sheet_pdf")
    print(f"[DEBUG] course_id param = {course_id}")
    print(f"[DEBUG] course_code param = {request.args.get('course_code')}")
    print(f"[DEBUG] Resolved Course ID = {course.id}")
    print(f"[DEBUG] Resolved Course Code = {course.course_code}")
    print(f"[DEBUG] Curriculum ID = {course.curriculum_id}")
    print(f"[DEBUG] Department = {course.curriculum.department.code}")
    print(f"[DEBUG] Regulation = {course.curriculum.regulation.name}")
    print(f"[DEBUG] Batch = {course.curriculum.batch.label}")
    print("=" * 80)

    schedule = ExamSchedule.query.filter_by(course_id=course.id).first()
    ay = AcademicYear.query.filter_by(is_current=True).first()
    existing = {}
    if schedule:
        for att in Attendance.query.filter_by(exam_schedule_id=schedule.id).all():
            existing[att.student_id] = att.status

    # ── STRICT CourseRegistration Query ──
    all_students = (
        db.session.query(Student)
        .join(CourseRegistration, CourseRegistration.student_id == Student.id)
        .filter(CourseRegistration.course_id == course.id)
        .order_by(Student.register_number)
        .all()
    )

    print(f"[DEBUG] Students Loaded = {len(all_students)}")
    for s in all_students[:20]:
        print(
            f"[DEBUG STUDENT] "
            f"{s.register_number} | "
            f"{s.department} | "
            f"{s.batch} | "
            f"{s.regulation}"
        )
    print("=" * 80)

    present_students = [s for s in all_students if existing.get(s.id, 'Present') == 'Present']
    total_present = len(present_students)
    BUNDLE_SIZE = 30
    num_bundles = max(1, math.ceil(total_present / BUNDLE_SIZE))
    sticker_map = {}
    if schedule:
        for ds in DummySticker.query.filter_by(exam_schedule_id=schedule.id).all():
            sticker_map[ds.student_id] = ds.dummy_number
    present_dummies = [sticker_map.get(s.id, '') for s in present_students]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=15*mm)
    NAVY, GOLD, WHITE, LIGHT = colors.HexColor('#1a2a5e'), colors.HexColor('#c9a227'), colors.white, colors.HexColor('#f5f5f5')
    styles = getSampleStyleSheet()
    ctr, sub = ParagraphStyle('C', fontName='Helvetica-Bold', fontSize=20, textColor=NAVY, alignment=TA_CENTER), ParagraphStyle('S', fontName='Helvetica-Bold', fontSize=11, textColor=NAVY, alignment=TA_CENTER)
    tiny, lbl, val2 = ParagraphStyle('T', fontName='Helvetica', fontSize=7, textColor=colors.grey, alignment=TA_CENTER), ParagraphStyle('L', fontName='Helvetica-Bold', fontSize=9, textColor=NAVY), ParagraphStyle('V2', fontName='Helvetica-Bold', fontSize=9, textColor=NAVY)
    val, foil, foil_v = ParagraphStyle('V', fontName='Helvetica-Bold', fontSize=9, textColor=GOLD), ParagraphStyle('F', fontName='Helvetica-Bold', fontSize=10, textColor=NAVY), ParagraphStyle('FV', fontName='Helvetica-Bold', fontSize=14, textColor=GOLD)
    page_w = A4[0] - 30*mm
    story = []

    for bundle_idx in range(num_bundles):
        # ── Header image
        get_institutional_header(story, page_w)
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(f'END SEMESTER EXAMINATIONS – {(ay.semester if ay else "EVEN").upper()} SEM {ay.label if ay else "2025-26"}', ParagraphStyle('ET', fontName='Helvetica-Bold', fontSize=10, textColor=NAVY, alignment=TA_CENTER)))
        story.append(Spacer(1, 3*mm))

        info1 = [[Paragraph('<b>BOARD</b>', lbl), Paragraph(':', lbl), Paragraph(course.curriculum.department.code, val2), Paragraph('<b>Course Code :</b>', lbl), Paragraph(f'<b>{course.course_code}</b>', val)]]
        story.append(Table(info1, colWidths=[page_w*0.12, page_w*0.03, page_w*0.32, page_w*0.25, page_w*0.28], style=[('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
        exam_str = f'{schedule.exam_date.strftime("%d.%m.%Y %a").upper()} – {schedule.session}' if schedule else 'Not Scheduled'
        info2 = [[Paragraph('<b>Exam Date &amp; Session :</b>', lbl), Paragraph(exam_str, val2), Paragraph('<b>Course Title :</b>', lbl), Paragraph(course.course_title, val2)]]
        story.append(Table(info2, colWidths=[page_w*0.27, page_w*0.25, page_w*0.18, page_w*0.30], style=[('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
        story.append(Spacer(1, 2*mm)); story.append(Paragraph('<b>Valuation Date &amp; Session :</b>', ParagraphStyle('VDL', fontName='Helvetica-Bold', fontSize=9, textColor=NAVY, alignment=TA_CENTER))); story.append(Spacer(1, 2*mm))

        dummy_data = [[Paragraph('<b>SNO</b>', ParagraphStyle('DH',fontName='Helvetica-Bold',fontSize=8,textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>DUMMY NO</b>', ParagraphStyle('DH2',fontName='Helvetica-Bold',fontSize=8,textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>SNO</b>', ParagraphStyle('DH3',fontName='Helvetica-Bold',fontSize=8,textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>DUMMY NO</b>', ParagraphStyle('DH4',fontName='Helvetica-Bold',fontSize=8,textColor=WHITE,alignment=TA_CENTER))]]
        # Calculate how many rows are needed for THIS bundle
        students_in_this_bundle = min(BUNDLE_SIZE, total_present - bundle_idx * BUNDLE_SIZE)
        rows_needed = math.ceil(students_in_this_bundle / 2)
        
        for row_i in range(rows_needed):
            l_idx = bundle_idx * BUNDLE_SIZE + row_i
            r_idx = bundle_idx * BUNDLE_SIZE + row_i + rows_needed
            l_sno = l_idx + 1
            r_sno = r_idx + 1
            # Left column dummy
            l_dummy = present_dummies[l_idx] if l_idx < (bundle_idx * BUNDLE_SIZE + students_in_this_bundle) else ''
            l_sno_str = str(l_sno) if l_dummy else ''
            # Right column dummy
            r_dummy = present_dummies[r_idx] if r_idx < (bundle_idx * BUNDLE_SIZE + students_in_this_bundle) else ''
            r_sno_str = str(r_sno) if r_dummy else ''
            dummy_data.append([
                Paragraph(l_sno_str, ParagraphStyle(f'LS{bundle_idx}_{row_i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph(f'<b>{l_dummy}</b>' if l_dummy else '', ParagraphStyle(f'LD{bundle_idx}_{row_i}', fontName='Helvetica-Bold', fontSize=8, textColor=NAVY, alignment=TA_CENTER)),
                Paragraph(r_sno_str, ParagraphStyle(f'RS{bundle_idx}_{row_i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph(f'<b>{r_dummy}</b>' if r_dummy else '', ParagraphStyle(f'RD{bundle_idx}_{row_i}', fontName='Helvetica-Bold', fontSize=8, textColor=NAVY, alignment=TA_CENTER)),
            ])


        dummy_tbl = Table(dummy_data, colWidths=[page_w*0.09, page_w*0.16, page_w*0.09, page_w*0.16]*2, style=[('BACKGROUND', (0,0), (-1,0), NAVY), ('TEXTCOLOR',  (0,0), (-1,0), WHITE), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')), ('LINEAFTER', (1,0), (1,-1), 1.5, NAVY)])
        foil_tbl = Table([[Paragraph('FOIL No.:', foil)], [Paragraph('', styles['Normal'])], [Paragraph('', styles['Normal'])], [Paragraph('', styles['Normal'])], [Paragraph('Packet No./Bundle No.:', foil)], [Paragraph(f'{bundle_idx+1} / {num_bundles}', foil_v)]], colWidths=[30*mm])
        story.append(Table([[dummy_tbl, foil_tbl]], colWidths=[page_w - 33*mm, 33*mm], style=[('VALIGN', (0,0), (-1,-1), 'TOP')]))
        story.append(Spacer(1, 5*mm)); story.append(HRFlowable(width='100%', thickness=0.5, color=NAVY))
        story.append(Table([[Paragraph('Name &amp; Signature of Examiner', ParagraphStyle('SE', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_LEFT)), Paragraph(f'Page {bundle_idx+1}', ParagraphStyle('PG', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_CENTER)), Paragraph('Name &amp; Signature of Chairman', ParagraphStyle('SC', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_RIGHT))]], colWidths=[page_w/3]*3))
        if bundle_idx < num_bundles - 1: story.append(PageBreak())

    doc.build(story); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'ESE_CoverSheet_{course.course_code}_{dt.date.today()}.pdf', mimetype='application/pdf')

@api.route('/ese/despatch-pdf', methods=['GET'])
@login_required
def ese_despatch_pdf():
    course_id = request.args.get('course_id')
    if not course_id:
        return jsonify({'message': 'course_id is MANDATORY for PDF generation'}), 400
        
    course = Course.query.get_or_404(course_id)

    print("=" * 80)
    print("[DEBUG] PDF/API ROUTE STARTED: ese_despatch_pdf")
    print(f"[DEBUG] course_id param = {course_id}")
    print(f"[DEBUG] course_code param = {request.args.get('course_code')}")
    print(f"[DEBUG] Resolved Course ID = {course.id}")
    print(f"[DEBUG] Resolved Course Code = {course.course_code}")
    print(f"[DEBUG] Curriculum ID = {course.curriculum_id}")
    print(f"[DEBUG] Department = {course.curriculum.department.code}")
    print(f"[DEBUG] Regulation = {course.curriculum.regulation.name}")
    print(f"[DEBUG] Batch = {course.curriculum.batch.label}")
    print("=" * 80)

    schedule = ExamSchedule.query.filter_by(course_id=course.id).first()
    ay = AcademicYear.query.filter_by(is_current=True).first()
    att_records = Attendance.query.filter_by(exam_schedule_id=schedule.id).all() if schedule else []
    existing = {att.student_id: att.status for att in att_records}

    # ── STRICT CourseRegistration Query ──
    all_students = (
        db.session.query(Student)
        .join(CourseRegistration, CourseRegistration.student_id == Student.id)
        .filter(CourseRegistration.course_id == course.id)
        .order_by(Student.register_number)
        .all()
    )

    print(f"[DEBUG] Students Loaded = {len(all_students)}")
    for s in all_students[:20]:
        print(
            f"[DEBUG STUDENT] "
            f"{s.register_number} | "
            f"{s.department} | "
            f"{s.batch} | "
            f"{s.regulation}"
        )
    print("=" * 80)

    present_list = [s for s in all_students if existing.get(s.id, 'Present') == 'Present']
    absent_list  = [s for s in all_students if existing.get(s.id) == 'Absent']
    mp_list      = [s for s in all_students if existing.get(s.id) == 'Malpractice']
    total, present_count, absent_count, malpractice_count = len(all_students), len(present_list), len(absent_list), len(mp_list)

    dept_codes = sorted(set(s.department for s in all_students if s.department))
    dept_str = ', '.join(dept_codes) if dept_codes else (course.curriculum.department.code if course.curriculum and course.curriculum.department else 'N/A')
    exam_date_disp = schedule.exam_date.strftime('%d.%m.%Y %a').upper() if schedule and schedule.exam_date else 'NOT SCHEDULED'
    session_disp   = (schedule.session or 'FN') if schedule else 'N/A'
    exam_period = "APRIL/MAY - 2026" if (ay and ay.semester and ay.semester.lower() == "even") else "NOV/DEC - 2025"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, 
        leftMargin=10*mm, rightMargin=10*mm, 
        topMargin=10*mm, bottomMargin=10*mm
    )
    
    NAVY, RED, LIGHT = colors.HexColor('#1a2a5e'), colors.HexColor('#dc2626'), colors.HexColor('#f3f4f6')
    styles = getSampleStyleSheet()
    def sty(name, **kw): return ParagraphStyle(name, **kw)
    
    title_sty = sty('T1', fontName='Helvetica-Bold', fontSize=10, textColor=colors.black, alignment=TA_CENTER, leading=12)
    meta_lbl  = sty('ML', fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.black)
    meta_val  = sty('MV', fontName='Helvetica-Bold', fontSize=8.5, textColor=RED)
    meta_val_b = sty('MVB', fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.black)
    table_hdr = sty('TH', fontName='Helvetica-Bold', fontSize=8, textColor=colors.black, alignment=TA_CENTER)
    cell_sty  = sty('CL', fontName='Helvetica', fontSize=7.5, textColor=colors.black)
    cell_sty_c = sty('CLC', fontName='Helvetica', fontSize=7.5, textColor=colors.black, alignment=TA_CENTER)
    sig_sty   = sty('SIG', fontName='Helvetica', fontSize=8.5, alignment=TA_CENTER)

    page_w = A4[0] - 20*mm
    story = []

    PAGE_SIZE = 28 # Reduced for safety with signatures
    total_pages = max(1, math.ceil(total / PAGE_SIZE))

    # Combine Absent and MP for the summary
    ab_mp_combined = []
    for s in absent_list:
        ab_mp_combined.append({'reg': s.register_number, 'name': s.name, 'type': 'AB'})
    for s in mp_list:
        ab_mp_combined.append({'reg': s.register_number, 'name': s.name, 'type': 'MP'})

    for page_idx in range(total_pages):
        # 1. Institutional Header
        get_institutional_header(story, page_w)
        
        # 2. Titles
        story.append(Paragraph(f'<b>END SEMESTER EXAMINATIONS - {exam_period}</b>', title_sty))
        story.append(Paragraph('<b>DESPATCH REPORT</b>', sty(f'DESP_{page_idx}', fontName='Helvetica-Bold', fontSize=12, alignment=TA_CENTER, spaceAfter=2)))
        story.append(HRFlowable(width='100%', thickness=0.5, color=colors.black))
        story.append(Spacer(1, 2*mm))

        # 3. Metadata Grid
        meta_data = [
            [
                Paragraph('<b>Department</b>', meta_lbl), Paragraph(':', meta_lbl), Paragraph(dept_str, meta_val_b),
                Paragraph('<b>Course Code</b>', meta_lbl), Paragraph(':', meta_lbl), Paragraph(course.course_code or 'N/A', meta_val)
            ],
            [
                Paragraph('<b>Exam Date & Session</b>', meta_lbl), Paragraph(':', meta_lbl), Paragraph(f"{exam_date_disp} – {session_disp}", meta_val_b),
                Paragraph('<b>Course Title</b>', meta_lbl), Paragraph(':', meta_lbl), Paragraph(course.course_title or 'N/A', meta_val_b)
            ],
            [
                Paragraph('<b>Total Strength</b>', meta_lbl), Paragraph(':', meta_lbl), Paragraph(str(total), meta_val_b),
                Paragraph('<b>No. of Present</b>', meta_lbl), Paragraph(':', meta_lbl), Paragraph(str(present_count), meta_val_b)
            ],
            [
                Paragraph('<b>No. of Absent</b>', meta_lbl), Paragraph(':', meta_lbl), Paragraph(str(absent_count), meta_val_b),
                Paragraph('<b>No. of Malpractice</b>', meta_lbl), Paragraph(':', meta_lbl), Paragraph(str(malpractice_count), meta_val_b)
            ]
        ]
        meta_tbl = Table(meta_data, colWidths=[page_w*0.18, page_w*0.02, page_w*0.35, page_w*0.18, page_w*0.02, page_w*0.25])
        meta_tbl.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 2),
        ]))
        story.append(meta_tbl)
        story.append(Spacer(1, 4*mm))

        # 4. Main Table
        main_hdr = [[Paragraph('<b>SNo</b>', table_hdr), Paragraph('<b>Register No.</b>', table_hdr), Paragraph('<b>Name of the Student</b>', table_hdr), Paragraph('<b>Answer Script No. / Remarks</b>', table_hdr)]]
        rows = []
        page_students = all_students[page_idx*PAGE_SIZE : (page_idx+1)*PAGE_SIZE]
        
        for i, s in enumerate(page_students):
            idx = page_idx * PAGE_SIZE + i + 1
            status = existing.get(s.id, 'Present')
            remarks = ''
            if status == 'Absent': 
                remarks = Paragraph('<b>ABSENT (AB)</b>', sty(f'RAB_{idx}', fontName='Helvetica-Bold', fontSize=8, textColor=RED, alignment=TA_CENTER))
            elif status == 'Malpractice': 
                remarks = Paragraph('<b>MALPRACTICE (MP)</b>', sty(f'RMP_{idx}', fontName='Helvetica-Bold', fontSize=8, textColor=RED, alignment=TA_CENTER))
            
            rows.append([Paragraph(str(idx), cell_sty_c), Paragraph(s.register_number or '', cell_sty_c), Paragraph(s.name or '', cell_sty), remarks])        

        main_tbl = Table(main_hdr + rows, colWidths=[page_w*0.08, page_w*0.22, page_w*0.35, page_w*0.35], repeatRows=1)
        main_tbl.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2.5),
            ('TOPPADDING', (0,0), (-1,-1), 2.5),
        ]))
        story.append(main_tbl)
        story.append(Spacer(1, 10*mm))
        
        # 5. Signatures
        sig_data = [[
            Paragraph('<b>__________________________</b><br/>Signature of the Hall Invigilator', sig_sty),
            Paragraph(f'Page {page_idx+1} of {total_pages}', sty(f'PG_{page_idx}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
            Paragraph('<b>__________________________</b><br/>Signature of the Chief Superintendent', sig_sty)
        ]]
        story.append(Table(sig_data, colWidths=[page_w/3]*3))

        if page_idx < total_pages - 1:
            story.append(PageBreak())

    # 6. Summary Page
    if ab_mp_combined:
        story.append(PageBreak())
        get_institutional_header(story, page_w)
        story.append(Paragraph('<b>LIST OF ABSENT / MALPRACTICE CANDIDATES</b>', sty('ABH', fontName='Helvetica-Bold', fontSize=10, alignment=TA_CENTER, spaceAfter=2)))
        ab_hdr = [[Paragraph('<b>Register No.</b>', table_hdr), Paragraph('<b>Name of the Student</b>', table_hdr), Paragraph('<b>Status</b>', table_hdr)]]
        ab_rows = []
        for item in ab_mp_combined:
            ab_rows.append([Paragraph(item['reg'] or '', cell_sty_c), Paragraph(item['name'] or '', cell_sty), Paragraph('ABSENT' if item['type']=='AB' else 'MALPRACTICE', cell_sty_c)])
        
        ab_tbl = Table(ab_hdr + ab_rows, colWidths=[page_w*0.3, page_w*0.4, page_w*0.3])
        ab_tbl.setStyle(TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('BACKGROUND', (0,0), (-1,0), LIGHT)]))
        story.append(ab_tbl)

    doc.build(story)
    buf.seek(0)
    filename = f'ESE_Despatch_{course.course_code}_{dt.date.today()}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')



@api.route('/ese/dummy-upload', methods=['POST'])
@login_required
def ese_dummy_upload():
    if current_user.role != 'admin': return jsonify({'message': 'Access denied'}), 403
    if 'file' not in request.files: return jsonify({'message': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith(('.xlsx', '.xls')): return jsonify({'message': 'Only Excel files accepted'}), 400
    try:
        wb = openpyxl.load_workbook(file, data_only=True, read_only=True); ws = wb.active
        added, skipped, errors = 0, 0, []
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row): continue
            regno, dummy_no, foil_no, course_code = str(row[0] or '').strip().upper(), str(row[1] or '').strip(), str(row[2] or '').strip(), str(row[3] or '').strip().upper()
            student = Student.query.filter(db.func.upper(Student.register_number) == regno).first()
            course = Course.query.filter(db.func.upper(Course.course_code) == course_code).first()
            schedule = ExamSchedule.query.filter_by(course_id=course.id).first() if course else None
            if student and schedule:
                ds = DummySticker.query.filter_by(student_id=student.id, exam_schedule_id=schedule.id).first()
                if ds: ds.dummy_number, ds.foil_number = dummy_no, foil_no
                else: db.session.add(DummySticker(student_id=student.id, exam_schedule_id=schedule.id, dummy_number=dummy_no, foil_number=foil_no))
                added += 1
            else: skipped += 1; errors.append(f"Row {row_num}: Student or Course not found")
        db.session.commit()
        return jsonify({'message': f'{added} record(s) saved, {skipped} skipped.', 'added': added, 'skipped': skipped, 'errors': errors[:20]})
    except Exception as e: db.session.rollback(); return jsonify({'message': f'Failed: {e}'}), 500

@api.route('/ese/dummy-template', methods=['GET'])
@login_required
def ese_dummy_template():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Dummy Numbers'
    cols = [('REGNO', 'e.g. 210825104001', '210825104001'), ('DUMMY NO', '7-digit number', '4981757'), ('FOIL NO', 'Foil number', '1473'), ('COURSE CODE', 'e.g. GE241203', 'GE241203')]
    for c, (l, h, s) in enumerate(cols, 1):
        ws.cell(1, c, l).fill = PatternFill('solid', fgColor='1A2A5E'); ws.cell(1, c, l).font = Font(bold=True, color='FFFFFF')
        ws.cell(2, c, h).fill = PatternFill('solid', fgColor='FFF9E6'); ws.cell(3, c, s)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='Dummy_Upload_Template.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@api.route('/ese/sticker-pdf', methods=['GET'])
@login_required
def ese_sticker_pdf():
    course_id = request.args.get('course_id')
    if not course_id:
        return jsonify({'message': 'course_id is MANDATORY for PDF generation'}), 400
        
    course = Course.query.get_or_404(course_id)

    print("=" * 80)
    print("[DEBUG] PDF/API ROUTE STARTED: ese_sticker_pdf")
    print(f"[DEBUG] course_id param = {course_id}")
    print(f"[DEBUG] course_code param = {request.args.get('course_code')}")
    print(f"[DEBUG] Resolved Course ID = {course.id}")
    print(f"[DEBUG] Resolved Course Code = {course.course_code}")
    print(f"[DEBUG] Curriculum ID = {course.curriculum_id}")
    print(f"[DEBUG] Department = {course.curriculum.department.code}")
    print(f"[DEBUG] Regulation = {course.curriculum.regulation.name}")
    print(f"[DEBUG] Batch = {course.curriculum.batch.label}")
    print("=" * 80)

    schedule = ExamSchedule.query.filter_by(course_id=course.id).first() if course else None
    if not schedule: return jsonify({'message': 'No schedule found'}), 404

    only_present = request.args.get('only_present') == 'true'
    present_ids = None
    if only_present:
        present_ids = [att.student_id for att in Attendance.query.filter_by(exam_schedule_id=schedule.id, status='Present').all()]
    stickers = {ds.student_id: ds for ds in DummySticker.query.filter_by(exam_schedule_id=schedule.id).all()}

    # ── STRICT CourseRegistration Query ──
    base_students = (
        db.session.query(Student)
        .join(CourseRegistration, CourseRegistration.student_id == Student.id)
        .filter(CourseRegistration.course_id == course.id)
        .order_by(Student.register_number)
        .all()
    )
    
    print(f"[DEBUG] Students Loaded = {len(base_students)}")
    for s in base_students[:20]:
        print(
            f"[DEBUG STUDENT] "
            f"{s.register_number} | "
            f"{s.department} | "
            f"{s.batch} | "
            f"{s.regulation}"
        )
    print("=" * 80)
    eligible = [s for s in base_students if (present_ids is None or s.id in present_ids) and s.id in stickers]
    if not eligible: return jsonify({'message': 'No eligible students found'}), 404

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, 
        leftMargin=7.6*mm, rightMargin=8.2*mm, 
        topMargin=1.9*mm, bottomMargin=0
    )
    NAVY = colors.HexColor('#1a2a5e')
    story = []
    
    # 12 students per page, 4 stickers per student row
    cell_w = (A4[0] - 15.8*mm) / 4
    cell_h = (A4[1] - 1.9*mm) / 12 # Exact 12 rows per page

    for page_idx in range(math.ceil(len(eligible)/12)):
        page_studs = eligible[page_idx*12 : (page_idx+1)*12]
        table_data = []
        for stu in page_studs:
            ds = stickers[stu.id]
            # Sticker 1: REGNO + DUMMY + COURSE
            # Stickers 2-4: DUMMY + COURSE
            row_cells = [
                Paragraph(f'<b>REG. NO : {stu.register_number}</b><br/><b>{ds.dummy_number}</b><br/><b>{course.course_code}</b>', ParagraphStyle('M', alignment=TA_CENTER, fontSize=8, leading=9, textColor=NAVY)),
                Paragraph(f'<b>{ds.dummy_number}</b><br/><b>{course.course_code}</b>', ParagraphStyle('C1', alignment=TA_CENTER, fontSize=9, leading=10)),
                Paragraph(f'<b>{ds.dummy_number}</b><br/><b>{course.course_code}</b>', ParagraphStyle('C2', alignment=TA_CENTER, fontSize=9, leading=10)),
                Paragraph(f'<b>{ds.dummy_number}</b><br/><b>{course.course_code}</b>', ParagraphStyle('C3', alignment=TA_CENTER, fontSize=9, leading=10))
            ]
            table_data.append(row_cells)
        
        tbl = Table(table_data, colWidths=[cell_w]*4, rowHeights=[cell_h]*len(page_studs))
        tbl.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('LINEAFTER', (0,0), (0,-1), 1.0, colors.grey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ('TOPPADDING', (0,0), (-1,-1), 2),
        ]))
        story.append(tbl)
        if page_idx < math.ceil(len(eligible)/12)-1:
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'Dummy_Stickers_{course.course_code}.pdf', mimetype='application/pdf')

@api.route('/results/my', methods=['GET'])
@login_required
def get_my_results():
    regno = request.args.get('regno') if current_user.role != 'student' else current_user.username
    student = Student.query.filter_by(register_number=regno).first()
    if not student: return jsonify({'message': 'Student not found'}), 404
    
    # NEW: Enforcement of Release status
    if current_user.role == 'student' and not student.result_published:
        return jsonify({
            'message': 'Your results have not been released yet. Please contact the COE office.',
            'status': 'not_released'
        }), 403
    results = []
    for s in DummySticker.query.filter_by(student_id=student.id).all():
        mark = FoilMark.query.filter_by(dummy_number=s.dummy_number, course_id=s.exam_schedule.course_id).first()
        if mark: results.append({'semester': s.exam_schedule.course.semester, 'course_code': s.exam_schedule.course.course_code, 'course_title': s.exam_schedule.course.course_title, 'grade': mark.grade or '-'})
    ay = AcademicYear.query.filter_by(is_current=True).first()
    return jsonify({'student_name': student.name, 'register_no': student.register_number, 'session_title': f"END SEMESTER EXAMINATION RESULTS - {ay.semester if ay else 'NOV/DEC'} {ay.label.split('-')[0] if ay else '2024'}", 'results': sorted(results, key=lambda x: (x['semester'], x['course_code']))})

@api.route('/ese/hallticket-pdf', methods=['GET'])
@login_required
def get_hallticket_pdf():
    ay = AcademicYear.query.filter_by(is_current=True).first()
    if not ay or not ay.hall_ticket_published:
        return jsonify({'message': 'Hall Tickets not yet released.'}), 403
    regno = current_user.username if current_user.role == 'student' else request.args.get('regno')
    student = Student.query.filter_by(register_number=regno).first()
    if not student: return jsonify({'message': 'Student not found'}), 404
    
    # Check Clearance Status
    clearance = FeeClearance.query.filter_by(student_id=student.id, academic_year_id=ay.id).first()
    if not clearance or not clearance.approved:
        return jsonify({
            'message': 'Your Hall Ticket is not yet approved. Please ensure your fees are paid and attendance is clear.',
            'status': 'not_cleared'
        }), 403
        
    schedules = ExamSchedule.query.join(Course).join(Curriculum).join(Department).filter(Department.code == student.department, Course.semester == student.semester, ExamSchedule.academic_year_id == ay.id).all()
    if not schedules: return jsonify({'message': 'No exam schedules found'}), 404
    schedules.sort(key=lambda x: x.exam_date if x.exam_date else dt.date.max)

    buf = io.BytesIO()
    def add_page_border(c, doc):
        c.saveState(); c.setStrokeColor(colors.black); c.setLineWidth(2); c.rect(20, 20, A4[0]-40, A4[1]-40)
        c.setLineWidth(0.5); c.rect(25, 25, A4[0]-50, A4[1]-50); c.restoreState()
    
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet(); story = []
    
    # ── Header image
    page_w = A4[0] - 80
    get_institutional_header(story, page_w)
    story.append(Spacer(1, 15))
    story.append(Paragraph(f"<b>END SEMESTER EXAMINATIONS – {ay.semester} {ay.label}</b>", ParagraphStyle('t1', alignment=1)))
    story.append(Spacer(1, 8)); story.append(Paragraph("<b>HALL TICKET</b>", ParagraphStyle('t2', alignment=1, fontSize=15, textColor=NAVY))); story.append(Spacer(1, 15))

    info_table = Table([["Register Number", student.register_number, "Semester", str(student.semester), ""], ["Name", student.name, "DOB", student.dob or "", "Photo"], ["Degree", f"{student.degree} {student.department}", "Reg", student.regulation, ""]], colWidths=[1.4*inch, 2.6*inch, 1.2*inch, 1.2*inch, 1*inch])
    info_table.setStyle(TableStyle([('GRID', (0,0), (-2,-1), 1.2, colors.black), ('GRID', (4,0), (4,2), 1.2, colors.black), ('SPAN', (4,0), (4,2)), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('FONTSIZE', (0,0), (-1,-1), 9)]))
    story.append(info_table); story.append(Spacer(1, 20))

    sched_rows = [[Paragraph("<b>Sem</b>", styles['Normal']), Paragraph("<b>Code</b>", styles['Normal']), Paragraph("<b>Title</b>", styles['Normal']), Paragraph("<b>Sem</b>", styles['Normal']), Paragraph("<b>Code</b>", styles['Normal']), Paragraph("<b>Title</b>", styles['Normal'])]]
    half = (len(schedules) + 1) // 2
    for i in range(half):
        s1 = schedules[i]; row = [str(s1.course.semester), s1.course.course_code, Paragraph(f"<font size=7>{s1.course.course_title}</font>", styles['Normal'])]
        if i+half < len(schedules): s2 = schedules[i+half]; row.extend([str(s2.course.semester), s2.course.course_code, Paragraph(f"<font size=7>{s2.course.course_title}</font>", styles['Normal'])])
        else: row.extend(["","",""])
        sched_rows.append(row)
    story.append(Table(sched_rows, colWidths=[0.4*inch, 0.7*inch, 2.5*inch]*2, style=[('GRID', (0,0), (-1,-1), 0.5, colors.black), ('FONTSIZE', (0,0), (-1,-1), 8)]))
    story.append(Spacer(1, 20))

    story.append(Paragraph("<b>NOTE:</b> 1. Instruction strictly followed. 2. Carry ID Card.", styles['Normal']))
    story.append(Spacer(1, 30))
    sig_table = Table([["Signature of Candidate", "Principal", "COE"]], colWidths=[2.4*inch]*3, style=[('ALIGN', (0,0), (-1,-1), 'CENTER'), ('LINEABOVE', (0,0), (-1,0), 0.5, colors.black)])
    story.append(sig_table)

    # 6. Instructions Page (Page 2)
    story.append(PageBreak())
    story.append(Paragraph("<b>INSTRUCTIONS TO THE CANDIDATE</b>", ParagraphStyle('inst_title', parent=styles['Normal'], alignment=1, fontSize=12)))
    story.append(Spacer(1, 15))
    
    inst_style = ParagraphStyle('inst', parent=styles['Normal'], fontSize=9, leading=13, alignment=4) # justified
    insts = [
        "1. Admission to the Examination is provisional.",
        "2. The Hall Ticket is issued subject to the candidate satisfying the Attendance and other requirements as per Rules, Regulations and Instructions by the Institution from time to time. If later, it is found that the candidate fails to comply with the above requirements, the examinations written by the candidate will be treated as cancelled.",
        "3. A seat marked with Register number will be provided to each candidate. Candidate will occupy the allotted seat at least 10 minutes before the commencement of the examination. In no case, candidate shall be allowed to occupy a seat other than the seat allotted to him.",
        "4. Normally the Candidate will not be permitted to enter the hall after the commencement of the examination. Only on extraordinary circumstances, the candidates will be permitted during the first thirty minutes of the examination after obtaining the written permission from the Principal / CS. Under any circumstances the Candidate shall not be permitted to enter the hall after the expiry of first thirty minutes.",
        "5. Candidate shall not be allowed to leave the examination hall before the expiry of 45 minutes from the commencement of examination. The candidate who leaves the examination hall during the period allotted for a paper will not be allowed to re-enter the hall within that period.",
        "6. Candidate who is suffering from infectious diseases of any kind shall not be admitted to the examination hall.",
        "7. Candidate is strictly prohibited from smoking inside the examination hall.",
        "8. Strict silence should be maintained in the examination hall.",
        "9. Candidate is required to bring his/her own pens, pencils and erasers. Candidate should use only blue or black ink while answering his/her papers.",
        "10. Before proceeding to answer the paper, the candidate should write his/her register number, semester, subject and date of the examination at the appropriate space provided in the first page of the answer book and nowhere else in the answer book or in any additional attachment like drawing sheet, smith chart etc.",
        "11. If a candidate writes his/her register number on any part of the answer book/sheets other than the one provided for or puts any special mark or writes anything which may disclose, in any way, the identity of the Candidate/College, he/she will render himself/herself liable for disciplinary action.",
        "12. Writing of wrong register number in the answer book will entail rejection of the answer book.",
        "13. Candidate is not allowed to exceed the prescribed time assigned to each paper.",
        "14. Candidate shall not talk/ask questions of any kind during the examination.",
        "15. Candidate shall not carry any written / printed matter, any paper material, cell phone, pen drive, ipad, programmable calculator, any unauthorized data sheet / table into the examination hall and if anything is found in his/her possession his/her shall be liable for disciplinary action.",
        "16. No Candidate shall pass any part or whole of answer papers or question papers to any other candidate. No candidate shall allow another candidate to copy from his/her answer paper or copy from the answer paper of another candidate. If found committing such malpractice, the involved candidates shall be liable for disciplinary action.",
        "17. Candidate found guilty of using unfair means of any nature shall be liable for disciplinary action.",
        "18. Candidate will have to hand over the answer book to the Invigilator / Chief Superintendent before leaving the examination hall.",
        "19. Candidate should produce the hall ticket on demand by the Invigilator / Chief Superintendent / Anna University Representative / Squad members.",
        "20. Candidate shall not write anything in the Hall Ticket.",
        "21. Candidate shall write only the Register No. in the space provided in the Question Paper. Any other writings in the Question Paper is prohibited and punishable."
    ]
    for i in insts:
        story.append(Paragraph(i, inst_style))
        story.append(Spacer(1, 5))

    story.append(Spacer(1, 20))
    story.append(Paragraph("<b>CONTROLLER OF EXAMINATIONS</b>", ParagraphStyle('coe', parent=styles['Normal'], alignment=2, fontSize=10)))

    doc.build(story, onFirstPage=add_page_border, onLaterPages=add_page_border); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'HallTicket_{regno}.pdf', mimetype='application/pdf')

@api.route('/ese/toggle-publish', methods=['POST'])
@login_required
def toggle_publish():
    if current_user.role != 'admin': return jsonify({'message': 'Access denied'}), 403
    
    data = request.get_json()
    target = data.get('target') # 'hall_ticket' or 'results'
    value  = data.get('value')
    
    ay = AcademicYear.query.filter_by(is_current=True).first()
    if not ay: return jsonify({'message': 'Active academic year not found'}), 404
    
    if target == 'hall_ticket': ay.hall_ticket_published = value
    elif target == 'results':  ay.results_published = value
    
    db.session.commit()
    audit_log.log("TOGGLE_PUBLISH", {"target": target, "value": value})
    return jsonify({'message': 'Status updated', 'hall_ticket_published': ay.hall_ticket_published, 'results_published': ay.results_published})

# ── Session & Day Wise Report PDF ─────────────────────────────────────
@api.route('/reports/session-daywise-pdf', methods=['GET'])
@login_required
def session_daywise_report_pdf():
    """Generate a detailed session & day wise exam report PDF"""
    date_str = request.args.get('date', '')
    if not date_str:
        date_str = dt.date.today().strftime('%Y-%m-%d')
    
    try:
        report_date = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        report_date = dt.date.today()

    ay = AcademicYear.query.filter_by(is_current=True).first()
    
    # Get all exam schedules for this date
    schedules = ExamSchedule.query.filter_by(exam_date=report_date).all()
    
    # Group by session
    fn_schedules = [s for s in schedules if s.session == 'FN']
    an_schedules = [s for s in schedules if s.session == 'AN']

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=15*mm)
    
    NAVY = colors.HexColor('#1a2a5e')
    GOLD = colors.HexColor('#c9a227')
    WHITE = colors.white
    LIGHT = colors.HexColor('#f5f5f5')
    
    styles = getSampleStyleSheet()
    page_w = A4[0] - 30*mm
    story = []

    # ── Header image
    get_institutional_header(story, page_w)
    story.append(Spacer(1, 2*mm))
    story.append(HRFlowable(width='100%', thickness=1.5, color=NAVY))
    story.append(Spacer(1, 2*mm))


    ay_label = f'{ay.semester.upper()} SEM {ay.label}' if ay else 'EVEN SEM 2025-26'
    story.append(Paragraph(f'END SEMESTER EXAMINATIONS – {ay_label}', ParagraphStyle('ET', fontName='Helvetica-Bold', fontSize=11, textColor=NAVY, alignment=TA_CENTER)))
    story.append(Paragraph('SESSION & DAY WISE REPORT', ParagraphStyle('SDR', fontName='Helvetica-Bold', fontSize=13, textColor=GOLD, alignment=TA_CENTER, spaceAfter=4)))
    story.append(HRFlowable(width='100%', thickness=0.5, color=NAVY))
    story.append(Spacer(1, 3*mm))

    # Date info
    story.append(Paragraph(f'<b>Date :</b> {report_date.strftime("%d.%m.%Y %A").upper()}', ParagraphStyle('DT', fontName='Helvetica-Bold', fontSize=10, textColor=NAVY)))
    story.append(Spacer(1, 5*mm))

    def build_session_table(session_label, session_schedules):
        """Build a table for one session (FN or AN)"""
        story.append(Paragraph(f'<b>SESSION : {session_label} ({"FORENOON" if session_label == "FN" else "AFTERNOON"})</b>', 
            ParagraphStyle(f'SH_{session_label}', fontName='Helvetica-Bold', fontSize=10, textColor=WHITE, backColor=NAVY)))
        story.append(Spacer(1, 2*mm))

        if not session_schedules:
            story.append(Paragraph('<i>No examinations scheduled for this session.</i>', 
                ParagraphStyle(f'NE_{session_label}', fontName='Helvetica-Oblique', fontSize=9, textColor=colors.grey, alignment=TA_CENTER)))
            story.append(Spacer(1, 5*mm))
            return

        # Table header
        header = [
            Paragraph('<b>S.No</b>', ParagraphStyle(f'TH1_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph('<b>Course Code</b>', ParagraphStyle(f'TH2_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph('<b>Course Title</b>', ParagraphStyle(f'TH3_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph('<b>Dept</b>', ParagraphStyle(f'TH4_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph('<b>Sem</b>', ParagraphStyle(f'TH5_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph('<b>Venue</b>', ParagraphStyle(f'TH6_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph('<b>Strength</b>', ParagraphStyle(f'TH7_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph('<b>Present</b>', ParagraphStyle(f'TH8_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
            Paragraph('<b>Absent</b>', ParagraphStyle(f'TH9_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)),
        ]
        table_data = [header]
        
        total_strength, total_present, total_absent = 0, 0, 0
        
        for i, sched in enumerate(session_schedules, 1):
            course = sched.course
            # Count students for this course's semester
            # Count only registered students for this specific course curriculum instance
            strength = CourseRegistration.query.filter_by(course_id=course.id).count()
            # Count attendance
            present = Attendance.query.filter_by(exam_schedule_id=sched.id, status='Present').count()
            absent = Attendance.query.filter_by(exam_schedule_id=sched.id, status='Absent').count()
            if present == 0 and absent == 0:
                present = strength  # default all present if no attendance yet
            
            total_strength += strength
            total_present += present
            total_absent += absent
            
            table_data.append([
                Paragraph(str(i), ParagraphStyle(f'SN{session_label}{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph(f'<b>{course.course_code}</b>', ParagraphStyle(f'CC{session_label}{i}', fontName='Helvetica-Bold', fontSize=8, textColor=NAVY, alignment=TA_CENTER)),
                Paragraph(course.course_title, ParagraphStyle(f'CT{session_label}{i}', fontName='Helvetica', fontSize=7.5)),
                Paragraph(course.curriculum.department.code, ParagraphStyle(f'DP{session_label}{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph(str(course.semester or '-'), ParagraphStyle(f'SM{session_label}{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph(sched.venue or 'MAIN HALL', ParagraphStyle(f'VN{session_label}{i}', fontName='Helvetica', fontSize=7.5, alignment=TA_CENTER)),
                Paragraph(f'<b>{strength}</b>', ParagraphStyle(f'ST{session_label}{i}', fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER)),
                Paragraph(str(present), ParagraphStyle(f'PR{session_label}{i}', fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#16a34a'), alignment=TA_CENTER)),
                Paragraph(str(absent), ParagraphStyle(f'AB{session_label}{i}', fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#dc2626'), alignment=TA_CENTER)),
            ])
        
        # Total row
        table_data.append([
            '', '', 
            Paragraph('<b>TOTAL</b>', ParagraphStyle(f'TOT_{session_label}', fontName='Helvetica-Bold', fontSize=8, textColor=NAVY)),
            '', '', '',
            Paragraph(f'<b>{total_strength}</b>', ParagraphStyle(f'TS_{session_label}', fontName='Helvetica-Bold', fontSize=9, textColor=NAVY, alignment=TA_CENTER)),
            Paragraph(f'<b>{total_present}</b>', ParagraphStyle(f'TP_{session_label}', fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#16a34a'), alignment=TA_CENTER)),
            Paragraph(f'<b>{total_absent}</b>', ParagraphStyle(f'TA_{session_label}', fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#dc2626'), alignment=TA_CENTER)),
        ])

        col_widths = [page_w*0.05, page_w*0.11, page_w*0.28, page_w*0.08, page_w*0.06, page_w*0.14, page_w*0.10, page_w*0.09, page_w*0.09]
        tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), NAVY),
            ('TEXTCOLOR', (0,0), (-1,0), WHITE),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('LINEBELOW', (0,0), (-1,0), 1.5, NAVY),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#f0f4f8')),
            ('LINEABOVE', (0,-1), (-1,-1), 1.5, NAVY),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 8*mm))

    # Build FN and AN tables
    build_session_table('FN', fn_schedules)
    build_session_table('AN', an_schedules)

    # Summary
    total_exams = len(schedules)
    story.append(HRFlowable(width='100%', thickness=0.5, color=NAVY))
    story.append(Spacer(1, 3*mm))
    story.append(Paragraph(f'Total Examinations on {report_date.strftime("%d.%m.%Y")} : <b>{total_exams}</b> | FN: <b>{len(fn_schedules)}</b> | AN: <b>{len(an_schedules)}</b>',
        ParagraphStyle('SUM', fontName='Helvetica', fontSize=9, textColor=NAVY, alignment=TA_CENTER)))
    story.append(Spacer(1, 20*mm))

    # Signatures
    sign_data = [['', '', ''], ['________________________', '________________________', '________________________'], ['Squad Member', 'Chief Superintendent', 'Controller of Examinations']]
    sig_t = Table(sign_data, colWidths=[page_w/3]*3)
    sig_t.setStyle(TableStyle([('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,2), (-1,2), 'Helvetica-Bold'), ('FONTSIZE', (0,2), (-1,2), 8), ('TEXTCOLOR', (0,2), (-1,2), NAVY)]))
    story.append(sig_t)

    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'Session_DayWise_Report_{report_date}.pdf', mimetype='application/pdf')

# ── QP Cover Report PDF ───────────────────────────────────────────────
@api.route('/reports/qp-cover-pdf', methods=['GET'])
@login_required
def qp_cover_report_pdf():
    """Generate QP Cover report PDF for the COE office"""
    date_str = request.args.get('date', dt.date.today().strftime('%Y-%m-%d'))
    try: report_date = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except: report_date = dt.date.today()

    ay = AcademicYear.query.filter_by(is_current=True).first()
    schedules = ExamSchedule.query.filter_by(exam_date=report_date).order_by(ExamSchedule.session, ExamSchedule.course_id).all()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=15*mm)
    NAVY, GOLD, WHITE = colors.HexColor('#1a2a5e'), colors.HexColor('#c9a227'), colors.white
    page_w = A4[0] - 30*mm
    story = []

    # ── Header image
    get_institutional_header(story, page_w)
    story.append(Spacer(1, 4*mm))
    story.append(Paragraph(f'<b>QP COVER ALLOTMENT REPORT – {report_date.strftime("%d.%m.%Y")}</b>', ParagraphStyle('T', alignment=TA_CENTER, fontSize=12, textColor=NAVY)))
    story.append(Spacer(1, 5*mm))

    data = [[Paragraph('<b>SNo</b>', ParagraphStyle('TH1', textColor=WHITE, alignment=TA_CENTER)), Paragraph('<b>Course</b>', ParagraphStyle('TH2', textColor=WHITE)), Paragraph('<b>Session</b>', ParagraphStyle('TH3', textColor=WHITE, alignment=TA_CENTER)), Paragraph('<b>Strength</b>', ParagraphStyle('TH4', textColor=WHITE, alignment=TA_CENTER)), Paragraph('<b>Packets Needed</b>', ParagraphStyle('TH5', textColor=WHITE, alignment=TA_CENTER)), Paragraph('<b>Issued By</b>', ParagraphStyle('TH6', textColor=WHITE, alignment=TA_CENTER))]]
    
    for i, s in enumerate(schedules, 1):
        strength = CourseRegistration.query.filter_by(course_id=s.course.id).count()
        packets = math.ceil(strength / 30)
        data.append([str(i), f"{s.course.course_code}\n{s.course.course_title}", s.session, str(strength), f"{packets} (of 30)", ""])

    tbl = Table(data, colWidths=[page_w*0.07, page_w*0.43, page_w*0.10, page_w*0.10, page_w*0.15, page_w*0.15])
    tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY), ('TEXTCOLOR',(0,0),(-1,0),WHITE), ('GRID',(0,0),(-1,-1),0.5,colors.grey), ('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('FONTSIZE',(0,0),(-1,-1),8)]))
    story.append(tbl)
    
    doc.build(story); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'QP_Cover_Report_{report_date}.pdf', mimetype='application/pdf')

# ── Attendance Status Report PDF ──────────────────────────────────────
@api.route('/reports/attendance-status-pdf', methods=['GET'])
@login_required
def attendance_status_report_pdf():
    """Consolidated Attendance Status for all halls"""
    date_str = request.args.get('date', dt.date.today().strftime('%Y-%m-%d'))
    try: report_date = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except: report_date = dt.date.today()

    schedules = ExamSchedule.query.filter_by(exam_date=report_date).all()
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=12*mm, rightMargin=12*mm, topMargin=10*mm, bottomMargin=10*mm)
    NAVY, WHITE = colors.HexColor('#1a2a5e'), colors.white
    page_w = A4[0] - 24*mm
    story = []

    # ── Header image
    get_institutional_header(story, page_w)
    story.append(Paragraph(f'<b>KEC EXAMINATION ATTENDANCE STATUS – {report_date.strftime("%d.%m.%Y")}</b>', ParagraphStyle('T', alignment=TA_CENTER, fontSize=12, textColor=NAVY)))
    story.append(Spacer(1, 5*mm))

    data = [[Paragraph('<b>SNo</b>', ParagraphStyle('TH',textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>Course</b>', ParagraphStyle('TH',textColor=WHITE)), Paragraph('<b>Sess</b>', ParagraphStyle('TH',textColor=WHITE)), Paragraph('<b>Total</b>', ParagraphStyle('TH',textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>Present</b>', ParagraphStyle('TH',textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>Absent</b>', ParagraphStyle('TH',textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>Status</b>', ParagraphStyle('TH',textColor=WHITE,alignment=TA_CENTER))]]
    
    for i, s in enumerate(schedules, 1):
        total = CourseRegistration.query.filter_by(course_id=s.course.id).count()
        present = Attendance.query.filter_by(exam_schedule_id=s.id, status='Present').count()
        absent = Attendance.query.filter_by(exam_schedule_id=s.id, status='Absent').count()
        status = "MARKED" if (present+absent) > 0 else "PENDING"
        color = colors.HexColor('#16a34a') if status == "MARKED" else colors.HexColor('#dc2626')
        
        data.append([str(i), f"{s.course.course_code}", s.session, str(total), str(present), str(absent), Paragraph(f'<b>{status}</b>', ParagraphStyle('ST', textColor=color, alignment=TA_CENTER, fontSize=8))])

    tbl = Table(data, colWidths=[page_w*0.07, page_w*0.35, page_w*0.08, page_w*0.12, page_w*0.12, page_w*0.12, page_w*0.14])
    tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY), ('TEXTCOLOR',(0,0),(-1,0),WHITE), ('GRID',(0,0),(-1,-1),0.5,colors.grey), ('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    story.append(tbl)
    
    doc.build(story); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'Attendance_Status_{report_date}.pdf', mimetype='application/pdf')

# ── Missing Attendance Report PDF ─────────────────────────────────────
@api.route('/reports/missing-attendance-pdf', methods=['GET'])
@login_required
def missing_attendance_report_pdf():
    """Identify halls that haven't submitted attendance yet"""
    date_str = request.args.get('date', dt.date.today().strftime('%Y-%m-%d'))
    try: report_date = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except: report_date = dt.date.today()

    # Schedules where no attendance records exist
    all_schedules = ExamSchedule.query.filter_by(exam_date=report_date).all()
    missing = []
    for s in all_schedules:
        if Attendance.query.filter_by(exam_schedule_id=s.id).count() == 0:
            missing.append(s)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=12*mm, bottomMargin=15*mm)
    NAVY, WHITE, RED = colors.HexColor('#1a2a5e'), colors.white, colors.HexColor('#dc2626')
    page_w = A4[0] - 30*mm
    story = []

    # ── Header image
    get_institutional_header(story, page_w)
    story.append(Paragraph('<b>MISSING ATTENDANCE RECORDS ALERT</b>', ParagraphStyle('T', alignment=TA_CENTER, fontSize=14, textColor=RED)))
    story.append(Paragraph(f'<b>Date: {report_date.strftime("%d.%m.%Y")}</b>', ParagraphStyle('D', alignment=TA_CENTER, fontSize=10)))
    story.append(Spacer(1, 6*mm))

    if not missing:
        story.append(Paragraph('<b>All examination halls have submitted their attendance records.</b>', ParagraphStyle('G', alignment=TA_CENTER, textColor=colors.HexColor('#16a34a'))))
    else:
        data = [[Paragraph('<b>SNo</b>', ParagraphStyle('TH',textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>Course Code</b>', ParagraphStyle('TH',textColor=WHITE)), Paragraph('<b>Course Title</b>', ParagraphStyle('TH',textColor=WHITE)), Paragraph('<b>Session</b>', ParagraphStyle('TH',textColor=WHITE,alignment=TA_CENTER)), Paragraph('<b>Venue / Hall</b>', ParagraphStyle('TH',textColor=WHITE,alignment=TA_CENTER))]]
        for i, s in enumerate(missing, 1):
            data.append([str(i), s.course.course_code, s.course.course_title, s.session, s.venue or 'NOT ASSIGNED'])
        
        tbl = Table(data, colWidths=[page_w*0.07, page_w*0.18, page_w*0.45, page_w*0.10, page_w*0.20])
        tbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY), ('TEXTCOLOR',(0,0),(-1,0),WHITE), ('GRID',(0,0),(-1,-1),0.5,colors.grey), ('VALIGN',(0,0),(-1,-1),'MIDDLE'), ('FONTSIZE',(0,0),(-1,-1),8)]))
        story.append(tbl)
    
    doc.build(story); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'Missing_Attendance_{report_date}.pdf', mimetype='application/pdf')

# -- Publication Visibility Controls ----------------------------------
@api.route('/ese/publication-status', methods=['GET'])
@login_required
def get_publication_status():
    ay = AcademicYear.query.filter_by(is_current=True).first()
    return jsonify({
        'hall_ticket_published': ay.hall_ticket_published if ay else False,
        'results_published': ay.results_published if ay else False
    })

@api.route('/ese/update-publication', methods=['POST'])
@login_required
def update_publication_status():
    if current_user.role != 'admin': return jsonify({'message': 'Admin only'}), 403
    d = request.get_json()
    target = d.get('target') # 'hall_ticket' or 'results'
    value = d.get('value', False)
    ay = AcademicYear.query.filter_by(is_current=True).first()
    if not ay: return jsonify({'message': 'Current Academic Year not set'}), 404
    if target == 'hall_ticket': ay.hall_ticket_published = value
    elif target == 'results': ay.results_published = value
    else: return jsonify({'message': 'Invalid target'}), 400
    db.session.commit()
    audit_log.log('UPDATE_PUBLICATION', {'target': target, 'value': value})
    return jsonify({'message': f'{target.title()} visibility updated', 'status': value})
