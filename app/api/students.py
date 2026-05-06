from flask import jsonify, request, send_file
from flask_login import login_required, current_user
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.models import Student, Degree, Department, Batch, Regulation, AcademicYear
from app import db, limiter
from app.api import api

@api.route('/students', methods=['GET'])
@login_required
def get_students():
    dept     = request.args.get('dept', '')
    batch    = request.args.get('batch', '')
    page     = request.args.get('page', type=int)
    per_page = request.args.get('per_page', 200, type=int)

    q = Student.query.order_by(Student.department, Student.name)
    if dept:  q = q.filter_by(department=dept)
    if batch: q = q.filter_by(batch=batch)

    if page:
        pag = q.paginate(page=page, per_page=per_page, error_out=False)
        students = pag.items
        total    = pag.total
    else:
        students = q.all()
        total    = len(students)

    return jsonify({
        'total': total,
        'students': [{
            'id': s.id, 'register_number': s.register_number,
            'name': s.name, 'department': s.department,
            'batch': s.batch, 'academic_year': s.academic_year,
            'semester': s.semester, 'degree': s.degree,
            'regulation': s.regulation, 'email': s.email or '',
            'phone': s.phone or '', 'dob': s.dob or '',
            'result_published': s.result_published
        } for s in students]
    })

@api.route('/students', methods=['POST'])
@login_required
def add_student():
    d = request.get_json()
    required = ['register_number','name','department','batch','academic_year']
    if not all(d.get(f) for f in required):
        return jsonify({'message': 'Required fields missing'}), 400
    if Student.query.filter_by(register_number=d['register_number'].strip().upper()).first():
        return jsonify({'message': 'Register number already exists'}), 409
    s = Student(
        register_number = d['register_number'].strip().upper(),
        name            = d['name'].strip(),
        department      = d['department'].strip(),
        batch           = d['batch'].strip(),
        academic_year   = int(d['academic_year']),
        semester        = int(d.get('semester') or 1),
        degree          = d.get('degree', 'BE'),
        regulation      = d.get('regulation', 'R2021'),
        email           = d.get('email', ''),
        phone           = d.get('phone', ''),
        dob             = d.get('dob', ''),
    )
    db.session.add(s)
    db.session.commit()
    return jsonify({'message': 'Student added', 'id': s.id}), 201

@api.route('/students/<int:sid>', methods=['DELETE'])
@login_required
def delete_student(sid):
    s = Student.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()
    return jsonify({'message': 'Deleted'})

@api.route('/students/template', methods=['GET'])
def student_excel_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    columns = [
        ('DEGREE',           'BE / B.Tech / ME / PhD',   'BE'),
        ('DEPARTMENT',       'AIDS / CSE / ECE / IT …',  'CSE'),
        ('BATCH',            '2021-2025 / 2022-2026 …',  '2021-2025'),
        ('REGULATION',       'R2021 / R2019 / R2017',    'R2021'),
        ('REGISTER NUMBER',  'e.g. 211CS001',            '211CS001'),
        ('FULL NAME',        'Student Full Name',        'Sample Student'),
        ('ACADEMIC YEAR',    '1 / 2 / 3 / 4',            '3'),
        ('SEMESTER',         '1 to 8',                   '6'),
        ('EMAIL',            'student@kec.ac.in',        'sample@kec.ac.in'),
        ('PHONE',            '10-digit mobile number',   '9876543210'),
        ('DATE OF BIRTH',    'YYYY-MM-DD',               '2003-05-15'),
    ]

    navy     = '0F1A3D'
    gold     = 'C9A227'
    hdr_fill = PatternFill('solid', fgColor=navy)
    sub_fill = PatternFill('solid', fgColor='F0F4FF')
    tip_fill = PatternFill('solid', fgColor='FFF9E6')

    for col, (label, _, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, (_, hint, _) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col, value=hint)
        cell.font      = Font(italic=True, color='7A6000', size=9, name='Calibri')
        cell.fill      = tip_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for col, (_, _, sample) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col, value=sample)
        cell.fill      = sub_fill
        cell.alignment = Alignment(horizontal='center')

    widths = [12, 14, 14, 12, 18, 24, 14, 12, 26, 16, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A3'

    ws2 = wb.create_sheet('Reference Values')
    refs = [
        ('A', 'DEGREE',      ['BE', 'B.Tech', 'ME', 'PhD']),
        ('C', 'DEPARTMENT',  ['AIDS','AIML','BME','CSE','ECE','IT','MECH','RAA']),
        ('E', 'BATCH',       ['2021-2025','2022-2026','2023-2027','2024-2028']),
        ('G', 'REGULATION',  ['R2021','R2019','R2017']),
        ('I', 'ACAD. YEAR',  ['1','2','3','4']),
        ('K', 'SEMESTER',    ['1','2','3','4','5','6','7','8']),
    ]
    for col, title, values in refs:
        cell = ws2[f'{col}1']
        cell.value      = title
        cell.font       = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        cell.fill       = hdr_fill
        cell.alignment  = Alignment(horizontal='center')
        ws2.column_dimensions[col].width = 14
        for i, v in enumerate(values, 2):
            c = ws2[f'{col}{i}']
            c.value     = v
            c.alignment = Alignment(horizontal='center')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='KEC_Student_Upload_Template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@api.route('/students/bulk', methods=['POST'])
@login_required
def bulk_upload_students():
    if current_user.role not in ['admin', 'coe']:
        logging.warning(f"Unauthorized student upload attempt by {current_user.username}")
        return jsonify({'message': 'Access denied'}), 403

    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded — use field name "file"'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'message': 'No file selected'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'message': 'Only .xlsx files are supported'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        logging.error(f"Student Excel read error: {e}")
        return jsonify({'message': f'Cannot read Excel file: {str(e)}'}), 400

    # Auto-detect column mapping by scanning first 2 rows
    col_map = {'deg': 0, 'dept': 1, 'bt': 2, 'reg_n': 3, 'reg': 4, 'name': 5, 'yr': 6, 'sem': 7, 'email': 8, 'phone': 9, 'dob': 10}
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        for i, val in enumerate(row):
            if not val: continue
            v = str(val).strip().upper()
            if 'NAME' in v and 'FATHER' not in v: col_map['name'] = i
            elif 'REGISTER NUMBER' in v or 'REG NO' in v or 'REGNO' in v: col_map['reg'] = i
            elif 'DEPARTMENT' in v or 'DEPT' in v or 'BRANCH' in v: col_map['dept'] = i
            elif 'DEGREE' in v: col_map['deg'] = i
            elif 'BATCH' in v: col_map['bt'] = i
            elif 'REGULATION' in v: col_map['reg_n'] = i
            elif 'YEAR' in v: col_map['yr'] = i
            elif 'SEM' in v: col_map['sem'] = i
            elif 'EMAIL' in v: col_map['email'] = i
            elif 'PHONE' in v or 'MOBILE' in v: col_map['phone'] = i
            elif 'DOB' in v or 'DATE OF BIRTH' in v: col_map['dob'] = i

    added, skipped, errors = 0, 0, []
    try:
        existing_regs = set(
            r[0] for r in db.session.execute(
                db.text("SELECT register_number FROM student")
            ).fetchall()
        )
    except Exception as e:
        logging.error(f"Error fetching existing regs: {e}")
        existing_regs = set()

    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue
        try:
            def get_v(key, def_val=''):
                idx = col_map[key]
                if idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
                return def_val

            deg   = get_v('deg', 'BE')
            dept  = get_v('dept').upper()
            bt    = get_v('bt')
            reg_n = get_v('reg_n', 'R2021')
            reg   = get_v('reg').upper()
            name  = get_v('name')
            try: yr = int(float(get_v('yr') or 1))
            except: yr = 1
            try: sem = int(float(get_v('sem') or 1))
            except: sem = 1
            email = get_v('email')
            phone = get_v('phone')
            dob   = get_v('dob')

            if not all([reg, name, dept, bt]):
                errors.append(f'Row {row_num}: Missing reg/name/dept/batch')
                continue

            if reg in existing_regs:
                skipped += 1
                continue

            db.session.add(Student(
                register_number=reg, name=name, department=dept,
                degree=deg, batch=bt, academic_year=yr,
                semester=sem, regulation=reg_n,
                email=email, phone=phone, dob=dob
            ))
            existing_regs.add(reg)
            added += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Row {row_num}: {str(e)}')

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'DB commit failed: {str(e)}', 'added': 0, 'skipped': skipped, 'errors': errors}), 500

    return jsonify({
        'message': f'{added} student(s) added, {skipped} skipped (duplicate reg no)',
        'added': added, 'skipped': skipped, 'errors': errors
    }), 201

from app.utils.logger import audit_log

@api.route('/students/<int:sid>/toggle-result', methods=['POST'])
@login_required
def toggle_student_result(sid):
    if current_user.role != 'admin': return jsonify({'message': 'Access denied'}), 403
    s = Student.query.get_or_404(sid)
    s.result_published = not s.result_published
    db.session.commit()
    audit_log.log("STUDENT_RESULT_TOGGLE", {"regno": s.register_number, "published": s.result_published})
    return jsonify({'message': 'Status updated', 'result_published': s.result_published})
