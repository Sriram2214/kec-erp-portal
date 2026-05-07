from flask import Blueprint, jsonify, request, send_file
import datetime
import os
import logging
from flask_login import login_user, logout_user, login_required, current_user
from app.models import (User, Student, Faculty, Course, ExamSchedule, Curriculum,
                        Degree, Department, Batch, Regulation, AcademicYear,
                        CourseAllocation, DummySticker, FoilMark)
from app import db, limiter
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

api = Blueprint('api', __name__, url_prefix='/api')

# ── Master Data ───────────────────────────────────────────────────────────────

@api.route('/master')
@login_required
def get_master():
    """Single call to fetch all master data for dropdowns."""
    return jsonify({
        'degrees':      [{'id': d.id, 'name': d.name} for d in Degree.query.all()],
        'departments':  [{'id': d.id, 'code': d.code, 'name': d.name, 'degree_id': d.degree_id}
                         for d in Department.query.order_by(Department.code).all()],
        'batches':      [{'id': b.id, 'label': b.label} for b in Batch.query.order_by(Batch.label).all()],
        'regulations':  [{'id': r.id, 'name': r.name} for r in Regulation.query.all()],
        'academic_years': [{'id': a.id, 'label': a.label, 'semester': a.semester, 'is_current': a.is_current}
                           for a in AcademicYear.query.all()],
    })

# ── Auth ─────────────────────────────────────────────────────────────────────

@api.route('/login', methods=['POST'])
def api_login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    user = User.query.filter(db.func.lower(User.username) == db.func.lower(username)).first()
    if user and user.check_password(password):
        login_user(user)
        return jsonify({'message': 'OK', 'user': {
            'id': user.id, 'username': user.username, 'role': user.role
        }})
    return jsonify({'message': 'Invalid credentials'}), 401

@api.route('/logout', methods=['POST'])
@login_required
def api_logout():
    logout_user()
    return jsonify({'message': 'Logged out'})

@api.route('/me')
@login_required
def api_me():
    return jsonify({'id': current_user.id, 'username': current_user.username, 'role': current_user.role})

# ── Dashboard ─────────────────────────────────────────────────────────────────

@api.route('/dashboard/stats')
@login_required
def dashboard_stats():
    return jsonify({
        'students':  Student.query.count(),
        'faculty':   Faculty.query.count(),
        'courses':   Course.query.count(),
        'schedules': ExamSchedule.query.count(),
    })

# ── Students ──────────────────────────────────────────────────────────────────

@api.route('/students', methods=['GET'])
@login_required
def get_students():
    # Support ?dept=CSE&batch=2021-2025&page=1&per_page=100
    dept     = request.args.get('dept', '')
    batch    = request.args.get('batch', '')
    page     = request.args.get('page', type=int)
    per_page = request.args.get('per_page', 200, type=int)

    q = Student.query.order_by(Student.department, Student.name)
    if dept:  q = q.filter_by(department=dept)
    if batch: q = q.filter_by(batch=batch)

    if page:
        pag = q.paginate(page=page, per_page=per_page, error_out=False)
        students = pag.items
        total    = pag.total
    else:
        students = q.all()
        total    = len(students)

    return jsonify({
        'total': total,
        'students': [{
            'id': s.id, 'register_number': s.register_number,
            'name': s.name, 'department': s.department,
            'batch': s.batch, 'academic_year': s.academic_year,
            'semester': s.semester, 'degree': s.degree,
            'regulation': s.regulation, 'email': s.email or '',
            'phone': s.phone or '', 'dob': s.dob or '',
        } for s in students]
    })

@api.route('/students', methods=['POST'])
@login_required
def add_student():
    d = request.get_json()
    required = ['register_number','name','department','batch','academic_year']
    if not all(d.get(f) for f in required):
        return jsonify({'message': 'Required fields missing'}), 400
    if Student.query.filter_by(register_number=d['register_number'].strip().upper()).first():
        return jsonify({'message': 'Register number already exists'}), 409
    s = Student(
        register_number = d['register_number'].strip().upper(),
        name            = d['name'].strip(),
        department      = d['department'].strip(),
        batch           = d['batch'].strip(),
        academic_year   = int(d['academic_year']),
        semester        = int(d.get('semester') or 1),
        degree          = d.get('degree', 'BE'),
        regulation      = d.get('regulation', 'R2021'),
        email           = d.get('email', ''),
        phone           = d.get('phone', ''),
        dob             = d.get('dob', ''),
    )
    db.session.add(s)
    db.session.commit()
    return jsonify({'message': 'Student added', 'id': s.id}), 201

@api.route('/students/<int:sid>', methods=['DELETE'])
@login_required
def delete_student(sid):
    s = Student.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()
    return jsonify({'message': 'Deleted'})

@api.route('/students/template', methods=['GET'])
@login_required
def student_excel_template():
    """Download Excel template matching the Individual Add form layout."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Students'

    # ── Columns in EXACT form order ──────────────────────────────────────────
    columns = [
        ('DEGREE',           'BE / B.Tech / ME / PhD',   'BE'),
        ('DEPARTMENT',       'AIDS / CSE / ECE / IT …',  'CSE'),
        ('BATCH',            '2021-2025 / 2022-2026 …',  '2021-2025'),
        ('REGULATION',       'R2021 / R2019 / R2017',    'R2021'),
        ('REGISTER NUMBER',  'e.g. 211CS001',            '211CS001'),
        ('FULL NAME',        'Student Full Name',        'Sample Student'),
        ('ACADEMIC YEAR',    '1 / 2 / 3 / 4',            '3'),
        ('SEMESTER',         '1 to 8',                   '6'),
        ('EMAIL',            'student@kec.ac.in',        'sample@kec.ac.in'),
        ('PHONE',            '10-digit mobile number',   '9876543210'),
        ('DATE OF BIRTH',    'YYYY-MM-DD',               '2003-05-15'),
    ]

    navy     = '0F1A3D'
    gold     = 'C9A227'
    hdr_fill = PatternFill('solid', fgColor=navy)
    sub_fill = PatternFill('solid', fgColor='F0F4FF')
    tip_fill = PatternFill('solid', fgColor='FFF9E6')

    # Row 1 — column header (navy bg, white bold)
    for col, (label, _, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font      = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
        cell.fill      = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 2 — hint/format tip (gold bg, italic)
    for col, (_, hint, _) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col, value=hint)
        cell.font      = Font(italic=True, color='7A6000', size=9, name='Calibri')
        cell.fill      = tip_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Row 3 — sample data row
    for col, (_, _, sample) in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col, value=sample)
        cell.fill      = sub_fill
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    widths = [12, 14, 14, 12, 18, 24, 14, 12, 26, 16, 18]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = 'A3'  # freeze header rows

    # ── Reference sheet ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Reference Values')
    refs = [
        ('A', 'DEGREE',      ['BE', 'B.Tech', 'ME', 'PhD']),
        ('C', 'DEPARTMENT',  ['AIDS','AIML','BME','CSE','ECE','IT','MECH','RAA']),
        ('E', 'BATCH',       ['2021-2025','2022-2026','2023-2027','2024-2028']),
        ('G', 'REGULATION',  ['R2021','R2019','R2017']),
        ('I', 'ACAD. YEAR',  ['1','2','3','4']),
        ('K', 'SEMESTER',    ['1','2','3','4','5','6','7','8']),
    ]
    for col, title, values in refs:
        cell = ws2[f'{col}1']
        cell.value      = title
        cell.font       = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        cell.fill       = hdr_fill
        cell.alignment  = Alignment(horizontal='center')
        ws2.column_dimensions[col].width = 14
        for i, v in enumerate(values, 2):
            c = ws2[f'{col}{i}']
            c.value     = v
            c.alignment = Alignment(horizontal='center')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='KCE_Student_Upload_Template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@api.route('/students/bulk', methods=['POST'])
@login_required
@limiter.limit("10 per hour")
def bulk_upload_students():
    """Bulk upload students from Excel file."""
    if current_user.role != 'admin':
        logging.warning(f"Unauthorized student upload attempt by {current_user.username}")
        return jsonify({'message': 'Access denied'}), 403

    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded — use field name "file"'}), 400
    file = request.files['file']
    if not file.filename:
        return jsonify({'message': 'No file selected'}), 400
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'message': 'Only .xlsx files are supported'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        logging.error(f"Student Excel read error: {e}")
        return jsonify({'message': f'Cannot read Excel file: {str(e)}'}), 400

    added, skipped, errors = 0, 0, []
    BATCH_SIZE = 500   # commit every 500 rows → handles 2000+ smoothly

    try:
        # Get existing reg numbers once (avoid N+1 queries)
        existing_regs = set(
            r[0] for r in db.session.execute(
                db.text("SELECT register_number FROM student")
            ).fetchall()
        )
    except Exception as e:
        logging.error(f"Error fetching existing regs: {e}")
        existing_regs = set()

    batch = []

    for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue
        try:
            deg   = str(row[0] or 'BE').strip()
            dept  = str(row[1] or '').strip().upper()
            bt    = str(row[2] or '').strip()
            reg_n = str(row[3] or 'R2021').strip()
            reg   = str(row[4] or '').strip().upper()
            name  = str(row[5] or '').strip()
            yr    = int(float(str(row[6] or 1)))
            sem   = int(float(str(row[7] or 1)))
            email = str(row[8] or '').strip()
            phone = str(row[9] or '').strip()
            dob   = str(row[10]).strip() if row[10] not in (None, '') else ''

            if not all([reg, name, dept, bt]):
                errors.append(f'Row {row_num}: Missing reg/name/dept/batch')
                continue

            if reg in existing_regs:
                skipped += 1
                continue

            batch.append(Student(
                register_number=reg, name=name, department=dept,
                degree=deg, batch=bt, academic_year=yr,
                semester=sem, regulation=reg_n,
                email=email, phone=phone, dob=dob
            ))
            existing_regs.add(reg)  # prevent duplicates within same file
            added += 1
        except Exception as e:
            db.session.rollback()
            errors.append(f'Row {row_num}: {str(e)}')

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'DB commit failed: {str(e)}', 'added': 0, 'skipped': skipped, 'errors': errors}), 500

    return jsonify({
        'message': f'{added} student(s) added, {skipped} skipped (duplicate reg no)',
        'added': added, 'skipped': skipped, 'errors': errors
    }), 201

# ── Faculty ───────────────────────────────────────────────────────────────────

@api.route('/faculty', methods=['GET'])
@login_required
def get_faculty():
    return jsonify([{
        'id': f.id, 'employee_id': f.employee_id,
        'name': f.name, 'department': f.department,
        'designation': f.designation or '',
        'email': f.email or '', 'phone': f.phone or '',
    } for f in Faculty.query.order_by(Faculty.department, Faculty.name).all()])

@api.route('/faculty', methods=['POST'])
@login_required
def add_faculty():
    d = request.get_json()
    if not all([d.get('employee_id'), d.get('name'), d.get('department')]):
        return jsonify({'message': 'All fields required'}), 400
    if Faculty.query.filter_by(employee_id=d['employee_id'].strip().upper()).first():
        return jsonify({'message': 'Employee ID already exists'}), 409
    f = Faculty(
        employee_id = d['employee_id'].strip().upper(),
        name        = d['name'].strip(),
        department  = d['department'].strip(),
        designation = d.get('designation', ''),
        email       = d.get('email', ''),
        phone       = d.get('phone', ''),
    )
    db.session.add(f)
    db.session.commit()
    return jsonify({'message': 'Faculty added', 'id': f.id}), 201

@api.route('/faculty/<int:fid>', methods=['DELETE'])
@login_required
def delete_faculty(fid):
    f = Faculty.query.get_or_404(fid)
    db.session.delete(f)
    db.session.commit()
    return jsonify({'message': 'Deleted'})

@api.route('/faculty/my-courses', methods=['GET'])
@login_required
def my_courses():
    if current_user.role != 'faculty':
        return jsonify([])
    
    faculty = Faculty.query.filter_by(employee_id=current_user.username).first()
    if not faculty:
        return jsonify([])
        
    allocations = CourseAllocation.query.filter_by(faculty_id=faculty.id).all()
    res = []
    for a in allocations:
        res.append({
            'allocation_id': a.id,
            'course_code': a.course.course_code,
            'course_title': a.course.course_title,
            'batch': a.batch,
            'section': a.section,
            'department': a.course.department
        })
    return jsonify(res)

# ── System Users (Staff accounts) ─────────────────────────────────────────────

@api.route('/staff', methods=['GET'])
@login_required
def get_staff():
    return jsonify([{
        'id': u.id, 'username': u.username, 'role': u.role
    } for u in User.query.order_by(User.username).all()])

@api.route('/staff', methods=['POST'])
@login_required
def add_staff():
    d = request.get_json()
    if not all([d.get('username'), d.get('password'), d.get('role')]):
        return jsonify({'message': 'All fields required'}), 400
    if User.query.filter_by(username=d['username'].strip()).first():
        return jsonify({'message': 'Username already exists'}), 409
    u = User(username=d['username'].strip(), role=d['role'])
    u.set_password(d['password'])
    db.session.add(u)
    db.session.commit()
    return jsonify({'message': 'User created', 'id': u.id}), 201

@api.route('/staff/<int:uid>', methods=['DELETE'])
@login_required
def delete_staff(uid):
    if uid == current_user.id:
        return jsonify({'message': 'Cannot delete yourself'}), 400
    u = User.query.get_or_404(uid)
    db.session.delete(u)
    db.session.commit()
    return jsonify({'message': 'Deleted'})

# ── Courses ───────────────────────────────────────────────────────────────────

@api.route('/courses', methods=['GET'])
@login_required
def get_courses():
    degree_id = request.args.get('degree_id')
    dept_id   = request.args.get('department_id')
    batch_id  = request.args.get('batch_id')
    reg_id    = request.args.get('regulation_id')
    
    q = Course.query.join(Curriculum)
    
    if degree_id: q = q.filter(Curriculum.degree_id == degree_id)
    if dept_id:   q = q.filter(Curriculum.department_id == dept_id)
    if batch_id:  q = q.filter(Curriculum.batch_id == batch_id)
    if reg_id:    q = q.filter(Curriculum.regulation_id == reg_id)
    
    courses = q.order_by(Course.semester, Course.course_code).all()
    
    return jsonify([{
        'id': c.id, 
        'course_code': c.course_code,
        'course_title': c.course_title,
        'semester': c.semester,
        'credits': c.credits,
        'is_lab': c.is_lab,
        'regulation': c.curriculum.regulation.name,
        'department': c.curriculum.department.code,
        'batch': c.curriculum.batch.label
    } for c in courses])

# ── Exam Schedule ─────────────────────────────────────────────────────────────

@api.route('/schedules', methods=['GET'])
@login_required
def get_schedules():
    return jsonify([{
        'id': s.id,
        'course_code':  s.course.course_code,
        'course_title': s.course.course_title,
        'exam_date':    s.exam_date.isoformat(),
        'session':      s.session,
    } for s in ExamSchedule.query.all()])

# ── ESE Attendance ────────────────────────────────────────────────────────────

@api.route('/ese/course-info', methods=['GET'])
@login_required
def ese_course_info():
    """Given ?course_code=GE241203, return course + exam schedule info."""
    code = request.args.get('course_code', '').strip().upper()
    if not code:
        return jsonify({'message': 'course_code required'}), 400

    course = Course.query.filter(
        db.func.upper(Course.course_code) == code
    ).first()
    if not course:
        return jsonify({'message': f'Course "{code}" not found'}), 404

    # Get exam schedule if exists
    schedule = ExamSchedule.query.filter_by(course_id=course.id).first()

    # Get current academic year
    ay = AcademicYear.query.filter_by(is_current=True).first()

    return jsonify({
        'course_code':  course.course_code,
        'course_title': course.course_title,
        'department':   course.department,
        'semester':     course.semester,
        'credits':      course.credits,
        'is_lab':       course.is_lab,
        'regulation':   course.regulation,
        'exam_date':    schedule.exam_date.strftime('%d.%m.%Y %a') if schedule else '',
        'session':      schedule.session if schedule else '',
        'academic_year': ay.label if ay else '',
    })


@api.route('/ese/students', methods=['GET'])
@login_required
def ese_students():
    """
    Get ALL students registered for a course code, across all departments.
    ?course_code=GE241203
    Returns students sorted by dept → register_number.
    """
    code = request.args.get('course_code', '').strip().upper()
    if not code:
        return jsonify({'message': 'course_code required'}), 400

    course = Course.query.filter(
        db.func.upper(Course.course_code) == code
    ).first()
    if not course:
        return jsonify({'message': f'Course "{code}" not found'}), 404

    # Get existing attendance for this course (for pre-filling)
    from app.models import Attendance, ExamSchedule as ES
    schedule = ES.query.filter_by(course_id=course.id).first()
    existing = {}
    if schedule:
        for att in Attendance.query.filter_by(exam_schedule_id=schedule.id).all():
            existing[att.student_id] = att.status

    # Fetch students by semester matching course's semester
    students = Student.query.filter_by(
        semester=course.semester
    ).order_by(Student.department, Student.register_number).all()

    # Fetch dummy stickers
    dummy_map = {}
    if schedule:
        for ds in DummySticker.query.filter_by(exam_schedule_id=schedule.id).all():
            dummy_map[ds.student_id] = ds.dummy_number

    return jsonify({
        'course_code':  course.course_code,
        'course_title': course.course_title,
        'schedule_id':  schedule.id if schedule else None,
        'students': [{
            'id':              s.id,
            'register_number': s.register_number,
            'name':            s.name,
            'department':      s.department,
            'batch':           s.batch,
            'status':          existing.get(s.id, 'Present'),  # default Present
            'dummy_number':    dummy_map.get(s.id),
        } for s in students]
    })


@api.route('/ese/attendance', methods=['POST'])
@login_required
def save_ese_attendance():
    """
    Save / update ESE attendance.
    Body: { course_code, exam_date, session, entries: [{student_id, status}] }
    """
    from app.models import Attendance, ExamSchedule as ES
    import datetime

    data = request.get_json()
    code    = data.get('course_code', '').strip().upper()
    entries = data.get('entries', [])

    course = Course.query.filter(
        db.func.upper(Course.course_code) == code
    ).first()
    if not course:
        return jsonify({'message': 'Course not found'}), 404

    # Upsert exam schedule if exam_date provided
    exam_date_str = data.get('exam_date', '')
    session_val   = data.get('session', 'FN')
    schedule = ES.query.filter_by(course_id=course.id).first()

    if not schedule and exam_date_str:
        try:
            edate = datetime.datetime.strptime(exam_date_str, '%Y-%m-%d').date()
        except Exception:
            edate = datetime.date.today()
        ay = AcademicYear.query.filter_by(is_current=True).first()
        schedule = ES(
            course_id        = course.id,
            exam_date        = edate,
            session          = session_val,
            academic_year_id = ay.id if ay else None,
        )
        db.session.add(schedule)
        db.session.flush()

    if not schedule:
        return jsonify({'message': 'No exam schedule found. Add exam date first.'}), 400

    # Upsert attendance records
    saved = 0
    for entry in entries:
        sid    = entry.get('student_id')
        status = entry.get('status', 'Present')
        att = Attendance.query.filter_by(
            student_id=sid, exam_schedule_id=schedule.id
        ).first()
        if att:
            att.status = status
        else:
            att = Attendance(student_id=sid, exam_schedule_id=schedule.id, status=status)
            db.session.add(att)
        saved += 1

    db.session.commit()
    return jsonify({'message': f'{saved} attendance record(s) saved.'})


@api.route('/ese/attendance-pdf', methods=['GET'])
@login_required
def ese_attendance_pdf():
    """
    Generate KCE-format ESE Attendance Sheet as PDF.
    ?course_code=GE241203
    """
    from app.models import Attendance, ExamSchedule as ES
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import os, datetime

    code = request.args.get('course_code', '').strip().upper()
    course = Course.query.filter(db.func.upper(Course.course_code) == code).first()
    if not course:
        return jsonify({'message': 'Course not found'}), 404

    schedule = ES.query.filter_by(course_id=course.id).first()
    ay = AcademicYear.query.filter_by(is_current=True).first()

    existing = {}
    if schedule:
        for att in Attendance.query.filter_by(exam_schedule_id=schedule.id).all():
            existing[att.student_id] = att.status

    students = Student.query.filter_by(
        semester=course.semester
    ).order_by(Student.department, Student.register_number).all()

    # ── Build PDF ──
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=12*mm, bottomMargin=12*mm,
    )

    NAVY  = colors.HexColor('#1a2a5e')
    GOLD  = colors.HexColor('#c9a227')
    WHITE = colors.white
    LIGHT = colors.HexColor('#f5f5f5')
    RED   = colors.HexColor('#dc2626')

    styles = getSampleStyleSheet()
    def sty(name, **kw):
        s = ParagraphStyle(name, **kw)
        return s

    hdr_title  = sty('HT', fontName='Helvetica-Bold', fontSize=18, textColor=NAVY,
                     alignment=TA_CENTER, spaceAfter=1)
    hdr_sub    = sty('HS', fontName='Helvetica-Bold', fontSize=10, textColor=NAVY,
                     alignment=TA_CENTER, spaceAfter=1)
    hdr_small  = sty('HSM', fontName='Helvetica', fontSize=7, textColor=colors.grey,
                     alignment=TA_CENTER, spaceAfter=2)
    sheet_title= sty('ST', fontName='Helvetica-Bold', fontSize=11, textColor=NAVY,
                     alignment=TA_CENTER, spaceAfter=4)
    label_sty  = sty('LB', fontName='Helvetica-Bold', fontSize=8.5, textColor=NAVY)
    val_sty    = sty('VL', fontName='Helvetica-Bold', fontSize=8.5, textColor=colors.HexColor('#c9a227'))
    val_black  = sty('VB', fontName='Helvetica-Bold', fontSize=8.5, textColor=NAVY)

    exam_date_disp = ''
    session_disp   = ''
    if schedule:
        exam_date_disp = schedule.exam_date.strftime('%d.%m.%Y %a').upper()
        session_disp   = schedule.session

    page_w = A4[0] - 30*mm  # usable width

    story = []

    from app.utils.pdf import get_institutional_header
    get_institutional_header(story, page_w)
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width='100%', thickness=1.5, color=NAVY))
    story.append(Spacer(1, 2*mm))

    ay_label = ay.label if ay else '2025-26'
    sem_label = ay.semester if ay else 'EVEN'
    story.append(Paragraph(
        f'END SEMESTER THEORY EXAMINATIONS – {sem_label.upper()} SEM {ay_label}',
        sheet_title
    ))
    story.append(Paragraph('ATTENDANCE SHEET', sty('ATT', fontName='Helvetica-Bold',
                            fontSize=13, textColor=NAVY, alignment=TA_CENTER, spaceAfter=4)))
    story.append(HRFlowable(width='100%', thickness=0.5, color=NAVY))
    story.append(Spacer(1, 3*mm))

    # ── Info Row ──
    dept_name = course.department
    info_data = [
        [
            Paragraph('<b>Degree &amp; Branch</b> :', label_sty),
            Paragraph(f'<font color="#1a2a5e"><b>B.E. / B.Tech – {dept_name}</b></font>', val_black),
            Paragraph('<b>Course Code :</b>', label_sty),
            Paragraph(f'<b>{course.course_code}</b>', val_sty),
        ],
        [
            Paragraph('<b>Exam Date &amp; Session :</b>', label_sty),
            Paragraph(f'<b>{exam_date_disp} – {session_disp}</b>' if exam_date_disp else '<i>Not Scheduled</i>', val_black),
            Paragraph('<b>Course Title :</b>', label_sty),
            Paragraph(f'<b>{course.course_title}</b>', val_black),
        ],
    ]
    info_tbl = Table(info_data, colWidths=[page_w*0.20, page_w*0.35, page_w*0.18, page_w*0.27])
    info_tbl.setStyle(TableStyle([
        ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 3*mm))

    # ── Attendance Table ──
    col_widths = [page_w*0.06, page_w*0.18, page_w*0.38, page_w*0.20, page_w*0.18]
    header_row = [
        Paragraph('<b>SNo</b>', sty('TH', fontName='Helvetica-Bold', fontSize=8,
                                    textColor=WHITE, alignment=TA_CENTER)),
        Paragraph('<b>Register No</b>', sty('TH2', fontName='Helvetica-Bold', fontSize=8,
                                             textColor=WHITE, alignment=TA_CENTER)),
        Paragraph('<b>Name of the Student</b>', sty('TH3', fontName='Helvetica-Bold', fontSize=8,
                                                     textColor=WHITE, alignment=TA_CENTER)),
        Paragraph('<b>"AB" for Absent</b>', sty('TH4', fontName='Helvetica-Bold', fontSize=8,
                                                 textColor=WHITE, alignment=TA_CENTER)),
        Paragraph('<b>Student Signature</b>', sty('TH5', fontName='Helvetica-Bold', fontSize=8,
                                                   textColor=WHITE, alignment=TA_CENTER)),
    ]

    table_data = [header_row]
    absent_count  = 0
    present_count = 0

    for i, s in enumerate(students, 1):
        status = existing.get(s.id, 'Present')
        is_absent = status in ('Absent', 'Malpractice')
        if is_absent:
            absent_count += 1
            ab_cell = Paragraph('<b>AB</b>', sty(f'AB{i}', fontName='Helvetica-Bold',
                                                  fontSize=9, textColor=RED, alignment=TA_CENTER))
        else:
            present_count += 1
            ab_cell = Paragraph('', styles['Normal'])

        row = [
            Paragraph(str(i), sty(f'SN{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
            Paragraph(s.register_number, sty(f'RN{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
            Paragraph(s.name, sty(f'NM{i}', fontName='Helvetica', fontSize=8)),
            ab_cell,
            Paragraph('', styles['Normal']),   # signature col — blank
        ]
        table_data.append(row)

    att_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    row_count = len(table_data)
    att_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0,0), (-1,0), NAVY),
        ('TEXTCOLOR',  (0,0), (-1,0), WHITE),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 8),
        ('ALIGN',      (0,0), (-1,0), 'CENTER'),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0),(-1,-1), 4),
        # Alternating rows
        *[('BACKGROUND', (0,r), (-1,r), LIGHT) for r in range(2, row_count, 2)],
        # Grid
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
        ('LINEBELOW',  (0,0), (-1,0), 1.5, NAVY),
        # Align reg no + sno center
        ('ALIGN',      (0,1), (1,-1), 'CENTER'),
        ('ALIGN',      (3,1), (4,-1), 'CENTER'),
    ]))
    story.append(att_table)
    story.append(Spacer(1, 5*mm))

    # ── Summary footer ──
    total = len(students)
    story.append(Paragraph(
        f'Total Students: <b>{total}</b> &nbsp;&nbsp;|&nbsp;&nbsp; '
        f'Present: <b>{present_count}</b> &nbsp;&nbsp;|&nbsp;&nbsp; '
        f'Absent: <b>{absent_count}</b>',
        sty('SUM', fontName='Helvetica', fontSize=8.5, textColor=NAVY, alignment=TA_CENTER)
    ))
    story.append(Spacer(1, 4*mm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=NAVY))

    # ── Sign row ──
    sign_data = [[
        Paragraph('Invigilator Signature', sty('SIG', fontName='Helvetica', fontSize=8,
                                                alignment=TA_CENTER, textColor=NAVY)),
        Paragraph('', styles['Normal']),
        Paragraph('Chief Superintendent', sty('SIG2', fontName='Helvetica', fontSize=8,
                                               alignment=TA_CENTER, textColor=NAVY)),
    ]]
    sign_tbl = Table(sign_data, colWidths=[page_w/3, page_w/3, page_w/3])
    sign_tbl.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(sign_tbl)

    doc.build(story)
    buf.seek(0)
    filename = f'ESE_Attendance_{code}_{datetime.date.today()}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/pdf')


# ── ESE Cover Sheet PDF ───────────────────────────────────────────────────────

@api.route('/ese/cover-sheet-pdf', methods=['GET'])
@login_required
def ese_cover_sheet_pdf():
    """
    Generate Answer-Bundle Cover Sheet PDF (KCE format).
    One page per bundle of 30 students.
    ?course_code=GE241203
    """
    from app.models import Attendance, ExamSchedule as ES
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable, KeepTogether)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import datetime, math, os

    code = request.args.get('course_code', '').strip().upper()
    course = Course.query.filter(db.func.upper(Course.course_code) == code).first()
    if not course:
        return jsonify({'message': 'Course not found'}), 404

    schedule = ES.query.filter_by(course_id=course.id).first()
    ay = AcademicYear.query.filter_by(is_current=True).first()

    # Only PRESENT students get an answer script
    existing = {}
    if schedule:
        for att in Attendance.query.filter_by(exam_schedule_id=schedule.id).all():
            existing[att.student_id] = att.status

    all_students = Student.query.filter_by(
        semester=course.semester
    ).order_by(Student.department, Student.register_number).all()

    present_students = [s for s in all_students if existing.get(s.id, 'Present') == 'Present']
    total_present = len(present_students)

    # Pagination — 30 per cover sheet page
    BUNDLE_SIZE = 30
    num_bundles = max(1, math.ceil(total_present / BUNDLE_SIZE))

    exam_date_disp = ''
    session_disp   = ''
    if schedule:
        exam_date_disp = schedule.exam_date.strftime('%d.%m.%Y %a').upper()
        session_disp   = schedule.session

    ay_label  = ay.label if ay else '2025-26'
    sem_label = (ay.semester if ay else 'EVEN').upper()

    # ── PDF Setup ──
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=12*mm, bottomMargin=15*mm,
    )

    NAVY  = colors.HexColor('#1a2a5e')
    GOLD  = colors.HexColor('#c9a227')
    WHITE = colors.white
    LIGHT = colors.HexColor('#f5f5f5')

    styles = getSampleStyleSheet()
    def sty(name, **kw):
        return ParagraphStyle(name, **kw)

    ctr  = sty('C', fontName='Helvetica-Bold', fontSize=20, textColor=NAVY, alignment=TA_CENTER)
    sub  = sty('S', fontName='Helvetica-Bold', fontSize=11, textColor=NAVY, alignment=TA_CENTER)
    tiny = sty('T', fontName='Helvetica',      fontSize=7,  textColor=colors.grey, alignment=TA_CENTER)
    lbl  = sty('L', fontName='Helvetica-Bold', fontSize=9,  textColor=NAVY)
    val  = sty('V', fontName='Helvetica-Bold', fontSize=9,  textColor=GOLD)
    val2 = sty('V2',fontName='Helvetica-Bold', fontSize=9,  textColor=NAVY)
    foil = sty('F', fontName='Helvetica-Bold', fontSize=10, textColor=NAVY)
    foil_v = sty('FV',fontName='Helvetica-Bold', fontSize=14, textColor=GOLD)
    page_w = A4[0] - 30*mm

    story = []

    for bundle_idx in range(num_bundles):
        start_idx = bundle_idx * BUNDLE_SIZE
        end_idx   = min(start_idx + BUNDLE_SIZE, total_present)
        bundle_students = present_students[start_idx:end_idx]

        # ── KCE Header ──
        story.append(Paragraph('KINGS', ctr))
        story.append(Paragraph('ENGINEERING COLLEGE', sub))
        story.append(Paragraph('AN AUTONOMOUS INSTITUTION', sty('A', fontName='Helvetica-Bold', fontSize=7.5, textColor=NAVY, alignment=TA_CENTER)))
        story.append(Paragraph('Accredited with NAAC and affiliated to Anna University', tiny))
        story.append(Paragraph('Chennai–Bangalore Highway, Irungattukottai, Sriperumbudur, Chennai – 602 117.', tiny))
        story.append(Paragraph('Ph.: 044-71224401-08, Fax: 044-71224410', tiny))
        story.append(Spacer(1, 3*mm))
        story.append(HRFlowable(width='100%', thickness=1.5, color=NAVY))
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(
            f'END SEMESTER EXAMINATIONS – {sem_label} SEM {ay_label}',
            sty('ET', fontName='Helvetica-Bold', fontSize=10, textColor=NAVY, alignment=TA_CENTER)
        ))
        story.append(Spacer(1, 3*mm))

        # ── Info rows ──
        info1 = [
            [Paragraph('<b>BOARD</b>', lbl),
             Paragraph(':',  lbl),
             Paragraph(course.department, val2),
             Paragraph('<b>Course Code :</b>', lbl),
             Paragraph(f'<b>{course.course_code}</b>', val)],
        ]
        info1_tbl = Table(info1, colWidths=[page_w*0.12, page_w*0.03,
                                             page_w*0.32, page_w*0.25, page_w*0.28])
        info1_tbl.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('TOPPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(info1_tbl)

        exam_str = f'{exam_date_disp} – {session_disp}' if exam_date_disp else 'Not Scheduled'
        info2 = [[
            Paragraph('<b>Exam Date &amp; Session :</b>', lbl),
            Paragraph(exam_str, val2),
            Paragraph('<b>Course Title :</b>', lbl),
            Paragraph(course.course_title, val2),
        ]]
        info2_tbl = Table(info2, colWidths=[page_w*0.27, page_w*0.25, page_w*0.18, page_w*0.30])
        info2_tbl.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
            ('BOTTOMPADDING',(0,0),(-1,-1),4),
            ('TOPPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(info2_tbl)
        story.append(Spacer(1, 2*mm))

        # Valuation date row (blank — to be filled manually)
        story.append(Paragraph(
            '<b>Valuation Date &amp; Session :</b>',
            sty('VDL', fontName='Helvetica-Bold', fontSize=9, textColor=NAVY, alignment=TA_CENTER)
        ))
        story.append(Spacer(1, 2*mm))

        # ── Dummy number table (2 column: 1-15, 16-30) ──
        col_w_half = page_w * 0.5
        dummy_header = [
            Paragraph('<b>SNO</b>',      sty('DH',fontName='Helvetica-Bold',fontSize=8,textColor=WHITE,alignment=TA_CENTER)),
            Paragraph('<b>DUMMY NO</b>', sty('DH2',fontName='Helvetica-Bold',fontSize=8,textColor=WHITE,alignment=TA_CENTER)),
            Paragraph('<b>SNO</b>',      sty('DH3',fontName='Helvetica-Bold',fontSize=8,textColor=WHITE,alignment=TA_CENTER)),
            Paragraph('<b>DUMMY NO</b>', sty('DH4',fontName='Helvetica-Bold',fontSize=8,textColor=WHITE,alignment=TA_CENTER)),
        ]
        dummy_data = [dummy_header]

        LEFT_HALF  = 15   # rows 1-15
        RIGHT_HALF = 15   # rows 16-30

        for row_i in range(LEFT_HALF):
            left_sno  = start_idx + row_i + 1
            right_sno = start_idx + row_i + LEFT_HALF + 1
            left_cell  = str(left_sno)  if left_sno  <= total_present else ''
            right_cell = str(right_sno) if right_sno <= total_present else ''
            dummy_data.append([
                Paragraph(left_cell,  sty(f'L{row_i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph('',         styles['Normal']),
                Paragraph(right_cell, sty(f'R{row_i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph('',         styles['Normal']),
            ])

        col_widths_d = [col_w_half*0.18, col_w_half*0.32,
                        col_w_half*0.18, col_w_half*0.32]
        dummy_tbl = Table(dummy_data, colWidths=col_widths_d, repeatRows=1)
        dummy_rows = len(dummy_data)
        dummy_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), NAVY),
            ('TEXTCOLOR',  (0,0), (-1,0), WHITE),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,0), 8),
            ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0),(-1,-1), 4),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('LINEBELOW',  (0,0), (-1,0), 1.5, NAVY),
            # Middle divider stronger
            ('LINEAFTER',  (1,0), (1,-1), 1.5, NAVY),
            *[('BACKGROUND', (0,r), (-1,r), LIGHT) for r in range(2, dummy_rows, 2)],
        ]))

        # FOIL No. / Bundle No. on right — put in a side table
        foil_data = [
            [Paragraph('FOIL No.:', foil)],
            [Paragraph('', styles['Normal'])],
            [Paragraph('', styles['Normal'])],
            [Paragraph('', styles['Normal'])],
            [Paragraph('Packet No./Bundle No.:', foil)],
            [Paragraph(f'{bundle_idx+1} / {num_bundles}', foil_v)],
        ]
        foil_tbl = Table(foil_data, colWidths=[30*mm])
        foil_tbl.setStyle(TableStyle([
            ('ALIGN',  (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))

        combined = Table(
            [[dummy_tbl, foil_tbl]],
            colWidths=[page_w - 33*mm, 33*mm]
        )
        combined.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(combined)
        story.append(Spacer(1, 5*mm))

        # ── Sign footer ──
        sign_data = [[
            Paragraph('Name &amp; Signature of Examiner',
                       sty('SE', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_LEFT)),
            Paragraph(f'Page {bundle_idx + 1}',
                       sty('PG', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_CENTER)),
            Paragraph('Name &amp; Signature of Chairman',
                       sty('SC', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_RIGHT)),
        ]]
        sign_tbl = Table(sign_data, colWidths=[page_w/3, page_w/3, page_w/3])
        sign_tbl.setStyle(TableStyle([
            ('TOPPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(HRFlowable(width='100%', thickness=0.5, color=NAVY))
        story.append(sign_tbl)

        # Page break between bundles
        if bundle_idx < num_bundles - 1:
            from reportlab.platypus import PageBreak
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    filename = f'ESE_CoverSheet_{code}_{datetime.date.today()}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/pdf')


# ── ESE Despatch PDF ──────────────────────────────────────────────────────────

@api.route('/ese/despatch-pdf', methods=['GET'])
@login_required
def ese_despatch_pdf():
    """
    Generate ESE Despatch PDF (KCE format).
    30 present students per page.
    AB/Malpractice summary shown in header only.
    ?course_code=GE241203
    """
    from app.models import Attendance, ExamSchedule as ES
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable, PageBreak)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import datetime, math

    code = request.args.get('course_code', '').strip().upper()
    course = Course.query.filter(db.func.upper(Course.course_code) == code).first()
    if not course:
        return jsonify({'message': 'Course not found'}), 404

    schedule = ES.query.filter_by(course_id=course.id).first()
    ay = AcademicYear.query.filter_by(is_current=True).first()

    existing = {}
    if schedule:
        for att in Attendance.query.filter_by(exam_schedule_id=schedule.id).all():
            existing[att.student_id] = att.status

    all_students = Student.query.filter_by(
        semester=course.semester
    ).order_by(Student.department, Student.register_number).all()

    total = len(all_students)
    present_list    = [s for s in all_students if existing.get(s.id, 'Present') == 'Present']
    absent_list     = [s for s in all_students if existing.get(s.id, 'Present') == 'Absent']
    malpractice_list= [s for s in all_students if existing.get(s.id, 'Present') == 'Malpractice']
    absent_count     = len(absent_list)
    malpractice_count= len(malpractice_list)
    present_count    = len(present_list)

    # All unique depts
    depts = sorted(set(s.department for s in all_students))
    dept_str = ', '.join(depts)

    exam_date_disp = ''
    session_disp   = ''
    if schedule:
        exam_date_disp = schedule.exam_date.strftime('%d.%m.%Y %a').upper()
        session_disp   = schedule.session

    ay_label  = ay.label if ay else '2025-26'
    sem_label = (ay.semester if ay else 'EVEN').upper()

    PAGE_SIZE   = 30
    total_pages = max(1, math.ceil(present_count / PAGE_SIZE))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15*mm, rightMargin=15*mm,
        topMargin=12*mm, bottomMargin=15*mm,
    )

    NAVY  = colors.HexColor('#1a2a5e')
    GOLD  = colors.HexColor('#c9a227')
    WHITE = colors.white
    LIGHT = colors.HexColor('#f5f5f5')
    RED   = colors.HexColor('#dc2626')

    styles = getSampleStyleSheet()
    def sty(name, **kw):
        return ParagraphStyle(name, **kw)

    page_w = A4[0] - 30*mm
    story  = []

    def build_page_header():
        """Returns story elements for KCE header + despatch info."""
        elems = []
        from app.utils.pdf import get_institutional_header
        get_institutional_header(elems, page_w)
        elems.append(Spacer(1, 3*mm))
        elems.append(HRFlowable(width='100%', thickness=1.5, color=NAVY))
        elems.append(Spacer(1, 1*mm))
        elems.append(Paragraph(
            f'END SEMESTER EXAMINATIONS – {sem_label} SEM {ay_label}',
            sty('DE', fontName='Helvetica-Bold', fontSize=10, textColor=NAVY, alignment=TA_CENTER)))
        elems.append(Paragraph('DESPATCH',
            sty('DESP', fontName='Helvetica-Bold', fontSize=13, textColor=NAVY, alignment=TA_CENTER)))
        elems.append(HRFlowable(width='100%', thickness=0.5, color=NAVY))
        elems.append(Spacer(1, 2*mm))

        lbl = sty('IL', fontName='Helvetica-Bold', fontSize=8.5, textColor=NAVY)
        val = sty('IV', fontName='Helvetica-Bold', fontSize=8.5, textColor=NAVY)
        vg  = sty('IVG',fontName='Helvetica-Bold', fontSize=8.5, textColor=GOLD)

        info1 = [[
            Paragraph('<b>Department</b>', lbl),
            Paragraph(':', lbl),
            Paragraph(dept_str, val),
            Paragraph('<b>Course Code :</b>', lbl),
            Paragraph(f'<b>{course.course_code}</b>', vg),
        ]]
        t1 = Table(info1, colWidths=[page_w*0.14,page_w*0.02,page_w*0.38,page_w*0.22,page_w*0.24])
        t1.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOTTOMPADDING',(0,0),(-1,-1),3),('TOPPADDING',(0,0),(-1,-1),3)]))
        elems.append(t1)

        exam_str = f'{exam_date_disp} – {session_disp}' if exam_date_disp else 'Not Scheduled'
        info2 = [[
            Paragraph('<b>Exam Date &amp; Session :</b>', lbl),
            Paragraph(exam_str, val),
            Paragraph('<b>Course Title :</b>', lbl),
            Paragraph(course.course_title, val),
        ]]
        t2 = Table(info2, colWidths=[page_w*0.26,page_w*0.28,page_w*0.16,page_w*0.30])
        t2.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOTTOMPADDING',(0,0),(-1,-1),3),('TOPPADDING',(0,0),(-1,-1),3)]))
        elems.append(t2)

        info3 = [[
            Paragraph(f'<b>Total No. of Students :</b> {total}', lbl),
            Paragraph(f'<b>No. of Malpractice :</b> {malpractice_count}', lbl),
        ],[
            Paragraph(f'<b>No. of Absent :</b> {absent_count}', lbl),
            Paragraph(f'<b>No. of Present :</b> {present_count}', lbl),
        ]]
        t3 = Table(info3, colWidths=[page_w*0.5, page_w*0.5])
        t3.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('BOTTOMPADDING',(0,0),(-1,-1),3),('TOPPADDING',(0,0),(-1,-1),3)]))
        elems.append(t3)
        elems.append(Spacer(1, 2*mm))
        return elems

    def build_sign_footer(page_num):
        sign_data = [[
            Paragraph('Name &amp; Signature of Examiner',
                       sty(f'SEX{page_num}', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_LEFT)),
            Paragraph(f'Page {page_num} of {total_pages}',
                       sty(f'PGN{page_num}', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_CENTER)),
            Paragraph('Name &amp; Signature of Chairman',
                       sty(f'SCH{page_num}', fontName='Helvetica', fontSize=8, textColor=NAVY, alignment=TA_RIGHT)),
        ]]
        sign_tbl = Table(sign_data, colWidths=[page_w/3, page_w/3, page_w/3])
        sign_tbl.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),6)]))
        return [HRFlowable(width='100%', thickness=0.5, color=NAVY), sign_tbl]

    # ── Table header builder ──
    def make_table_header():
        th = sty('TH', fontName='Helvetica-Bold', fontSize=8, textColor=WHITE, alignment=TA_CENTER)
        return [
            Paragraph('<b>SNo</b>', th),
            Paragraph('<b>Register No.</b>', th),
            Paragraph('<b>Name of the Student</b>', th),
        ]

    col_widths_p = [page_w*0.07, page_w*0.20, page_w*0.73]

    # ── Build pages ──
    for page_idx in range(total_pages):
        story += build_page_header()

        start = page_idx * PAGE_SIZE
        end   = min(start + PAGE_SIZE, present_count)
        page_students = present_list[start:end]

        table_data = [make_table_header()]
        for i, s in enumerate(page_students, start + 1):
            bg = LIGHT if i % 2 == 0 else WHITE
            row = [
                Paragraph(str(i), sty(f'SN{page_idx}_{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph(s.register_number, sty(f'RN{page_idx}_{i}', fontName='Helvetica', fontSize=8, alignment=TA_CENTER)),
                Paragraph(s.name, sty(f'NM{page_idx}_{i}', fontName='Helvetica', fontSize=8)),
            ]
            table_data.append(row)

        n_rows = len(table_data)
        tbl = Table(table_data, colWidths=col_widths_p, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), NAVY),
            ('TEXTCOLOR',  (0,0), (-1,0), WHITE),
            ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0), (-1,0), 8),
            ('ALIGN',      (0,0), (-1,0), 'CENTER'),
            ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0),(-1,-1), 4),
            ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('LINEBELOW',  (0,0), (-1,0), 1.5, NAVY),
            ('ALIGN',      (0,1), (1,-1), 'CENTER'),
            *[('BACKGROUND', (0,r), (-1,r), LIGHT) for r in range(2, n_rows, 2)],
        ]))
        story.append(tbl)
        story.append(Spacer(1, 4*mm))

        story += build_sign_footer(page_idx + 1)

        if page_idx < total_pages - 1:
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    filename = f'ESE_Despatch_{code}_{datetime.date.today()}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/pdf')


# ── Dummy Number Upload ───────────────────────────────────────────────────────

@api.route('/ese/dummy-upload', methods=['POST'])
@login_required
def ese_dummy_upload():
    """
    Upload Excel with dummy numbers.
    Expected columns: REGNO | DUMMY NO | FOIL NO | COURSE CODE
    Rows start from row 2 (row 1 = header).
    """
    if current_user.role != 'admin':
        logging.warning(f"Unauthorized dummy sticker upload attempt by {current_user.username}")
        return jsonify({'message': 'Access denied'}), 403

    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
    file = request.files['file']
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'message': 'Only Excel files accepted'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
        ws = wb.active
    except Exception as e:
        logging.error(f"DummySticker Excel read error: {e}")
        return jsonify({'message': f'Cannot read file: {e}'}), 400

    added, skipped, errors = 0, 0, []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(c for c in row if c is not None):
            continue
        try:
            regno      = str(row[0] or '').strip().upper()
            dummy_no   = str(row[1] or '').strip()
            foil_no    = str(row[2] or '').strip()
            course_code= str(row[3] or '').strip().upper()

            if not all([regno, dummy_no, foil_no, course_code]):
                errors.append(f'Row {row_num}: Missing fields')
                continue

            student = Student.query.filter(
                db.func.upper(Student.register_number) == regno
            ).first()
            if not student:
                errors.append(f'Row {row_num}: Student {regno} not found')
                skipped += 1
                continue

            course = Course.query.filter(
                db.func.upper(Course.course_code) == course_code
            ).first()
            if not course:
                errors.append(f'Row {row_num}: Course {course_code} not found')
                skipped += 1
                continue

            schedule = ExamSchedule.query.filter_by(course_id=course.id).first()
            if not schedule:
                errors.append(f'Row {row_num}: No exam schedule for {course_code}')
                skipped += 1
                continue

            # Upsert
            existing = DummySticker.query.filter_by(
                student_id=student.id, exam_schedule_id=schedule.id
            ).first()
            if existing:
                existing.dummy_number = dummy_no
                existing.foil_number  = foil_no
            else:
                ds = DummySticker(
                    student_id       = student.id,
                    exam_schedule_id = schedule.id,
                    dummy_number     = dummy_no,
                    foil_number      = foil_no,
                )
                db.session.add(ds)
            added += 1
        except Exception as e:
            errors.append(f'Row {row_num}: {e}')

    try:
        db.session.commit()
        logging.info(f"ADMIN {current_user.username} uploaded {added} dummy stickers successfully")
        return jsonify({
            'message': f'{added} record(s) saved, {skipped} skipped.',
            'added': added, 'skipped': skipped, 'errors': errors[:20]
        })
    except Exception as e:
        db.session.rollback()
        logging.error(f"Dummy Sticker DB commit error: {e}")
        return jsonify({'message': 'Failed to save dummy stickers to database'}), 500


@api.route('/ese/dummy-template', methods=['GET'])
@login_required
def ese_dummy_template():
    """Download Excel template for dummy number upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dummy Numbers'

    navy     = '1A2A5E'
    gold     = 'C9A227'
    hdr_fill = PatternFill('solid', fgColor=navy)
    tip_fill = PatternFill('solid', fgColor='FFF9E6')

    cols = [
        ('REGNO',        'e.g. 210825104001', '210825104001'),
        ('DUMMY NO',     '7-digit number',    '4981757'),
        ('FOIL NO',      'Foil number',       '1473'),
        ('COURSE CODE',  'e.g. GE241203',     'GE241203'),
    ]
    for c, (label, hint, sample) in enumerate(cols, 1):
        h = ws.cell(row=1, column=c, value=label)
        h.font      = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
        h.fill      = hdr_fill
        h.alignment = Alignment(horizontal='center')

        t = ws.cell(row=2, column=c, value=hint)
        t.font      = Font(italic=True, color='7A6000', size=9, name='Calibri')
        t.fill      = tip_fill
        t.alignment = Alignment(horizontal='center')

        s = ws.cell(row=3, column=c, value=sample)
        s.alignment = Alignment(horizontal='center')

    for c, w in enumerate([18, 14, 14, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name='Dummy_Upload_Template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Courses by Exam Date ──────────────────────────────────────────────────────

@api.route('/ese/courses-by-date', methods=['GET'])
@login_required
def ese_courses_by_date():
    """
    Get list of courses scheduled on a given exam date.
    ?date=2026-04-20
    Returns courses with sticker availability flag.
    """
    import datetime as dt

    date_str = request.args.get('date', '').strip()
    if not date_str:
        return jsonify({'message': 'date required (YYYY-MM-DD)'}), 400

    try:
        exam_date = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'message': 'Invalid date format'}), 400

    schedules = ExamSchedule.query.filter_by(exam_date=exam_date).all()
    result = []
    for sch in schedules:
        c = sch.course
        sticker_count = DummySticker.query.filter_by(exam_schedule_id=sch.id).count()
        # Count present students
        from app.models import Attendance
        present_count = Attendance.query.filter_by(
            exam_schedule_id=sch.id, status='Present'
        ).count()
        result.append({
            'schedule_id':   sch.id,
            'course_code':   c.course_code,
            'course_title':  c.course_title,
            'department':    c.department,
            'session':       sch.session,
            'sticker_count': sticker_count,
            'present_count': present_count,
        })

    result.sort(key=lambda x: x['course_code'])
    return jsonify(result)


# ── Dummy Sticker PDF ─────────────────────────────────────────────────────────

@api.route('/ese/sticker-pdf', methods=['GET'])
@login_required
def ese_sticker_pdf():
    """
    Generate Dummy Sticker PDF.
    12 students per page (4 stickers × 3 rows).
    Each row = 1 student × 4 sticker copies.
    Sticker 1: REGNO + DUMMY NO + COURSE CODE (master)
    Stickers 2-4: DUMMY NO + COURSE CODE only
    Excludes Absent and Malpractice students.
    ?course_code=GE241203
    """
    from app.models import Attendance
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm, inch
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, PageBreak)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    import datetime, math

    code = request.args.get('course_code', '').strip().upper()
    course = Course.query.filter(db.func.upper(Course.course_code) == code).first()
    if not course:
        return jsonify({'message': 'Course not found'}), 404

    schedule = ExamSchedule.query.filter_by(course_id=course.id).first()
    if not schedule:
        return jsonify({'message': 'No exam schedule found for this course'}), 404

    # Get present students only
    present_ids = set(
        a.student_id for a in
        Attendance.query.filter_by(exam_schedule_id=schedule.id, status='Present').all()
    )
    # If no attendance recorded yet, treat all as present
    att_count = Attendance.query.filter_by(exam_schedule_id=schedule.id).count()
    if att_count == 0:
        present_ids = None  # all students

    # Get dummy stickers for this schedule
    stickers = DummySticker.query.filter_by(exam_schedule_id=schedule.id).all()
    sticker_map = {ds.student_id: ds for ds in stickers}

    all_students = Student.query.filter_by(
        semester=course.semester
    ).order_by(Student.department, Student.register_number).all()

    # Filter: only present, and must have dummy sticker
    if present_ids is None:
        eligible = [s for s in all_students if s.id in sticker_map]
    else:
        eligible = [s for s in all_students if s.id in present_ids and s.id in sticker_map]

    if not eligible:
        return jsonify({'message': 'No eligible students with dummy numbers. Upload dummy numbers first.'}), 404

    # ── PDF Setup — label-sheet style ──
    buf = io.BytesIO()
    PAGE_W, PAGE_H = A4   # 595.27, 841.89 pts

    LEFT_M  = 7.6*mm
    RIGHT_M = 8.2*mm
    TOP_M   = 1.9*mm
    BOT_M   = 0*mm

    usable_w = PAGE_W - LEFT_M - RIGHT_M
    usable_h = PAGE_H - TOP_M - BOT_M

    COLS_PER_ROW = 4
    ROWS_PER_PAGE = 3   # 4×3 = 12 students per page
    STUDS_PER_PAGE = COLS_PER_ROW * ROWS_PER_PAGE  # 12

    cell_w = usable_w / COLS_PER_ROW
    cell_h = usable_h / ROWS_PER_PAGE

    NAVY  = colors.HexColor('#1a2a5e')
    GOLD  = colors.HexColor('#c9a227')
    BLACK = colors.black
    WHITE = colors.white
    LGREY = colors.HexColor('#f5f5f5')

    def cell_sty(name, **kw):
        return ParagraphStyle(name, **kw)

    regno_sty  = cell_sty('REG',  fontName='Helvetica-Bold', fontSize=8.5,
                           textColor=NAVY, alignment=TA_CENTER, leading=12)
    dummy_sty  = cell_sty('DMY',  fontName='Helvetica-Bold', fontSize=10,
                           textColor=BLACK, alignment=TA_CENTER, leading=14)
    course_sty = cell_sty('CRS',  fontName='Helvetica-Bold', fontSize=8.5,
                           textColor=NAVY, alignment=TA_CENTER, leading=12)
    dummy_only = cell_sty('DMYO', fontName='Helvetica-Bold', fontSize=10,
                           textColor=BLACK, alignment=TA_CENTER, leading=14)

    def make_master_cell(student, ds):
        """First sticker — includes REGNO."""
        content = (
            f'<b>REGNO : {student.register_number}</b><br/>'
            f'<b>{ds.dummy_number}</b><br/>'
            f'<b>{course.course_code}</b>'
        )
        return Paragraph(content, cell_sty(
            f'M{student.id}', fontName='Helvetica-Bold', fontSize=8.5,
            textColor=NAVY, alignment=TA_CENTER, leading=13
        ))

    def make_copy_cell(ds):
        """Copy stickers — DUMMY NO + COURSE only."""
        content = (
            f'<b>{ds.dummy_number}</b><br/>'
            f'<b>{course.course_code}</b>'
        )
        return Paragraph(content, cell_sty(
            f'C{ds.id}x', fontName='Helvetica-Bold', fontSize=9.5,
            textColor=BLACK, alignment=TA_CENTER, leading=14
        ))

    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=LEFT_M, rightMargin=RIGHT_M,
        topMargin=TOP_M, bottomMargin=BOT_M,
    )

    story = []
    total_pages = max(1, math.ceil(len(eligible) / STUDS_PER_PAGE))

    for page_idx in range(total_pages):
        page_start = page_idx * STUDS_PER_PAGE
        page_studs = eligible[page_start: page_start + STUDS_PER_PAGE]

        # Build 3 rows of 4 stickers each
        for row_idx in range(ROWS_PER_PAGE):
            row_start = row_idx * COLS_PER_ROW
            row_studs = page_studs[row_start: row_start + COLS_PER_ROW]

            # Each sticker = 4 cells: master + 3 copies
            # 4 students in a visual row → 4 groups of 4 cells
            # But each STUDENT occupies 4 cells in a row
            # Visually: [STU1-master][STU1-copy][STU1-copy][STU1-copy]
            # So this is 1 student per row of 4 columns
            # For 4 students across we'd need 4 rows, but image shows 1 student = 1 row
            # Going with: each row = 1 student × 4 cells

            for stu in row_studs:
                ds = sticker_map[stu.id]
                row_cells = [
                    make_master_cell(stu, ds),   # col 1: REGNO + DUMMY + COURSE
                    make_copy_cell(ds),           # col 2: DUMMY + COURSE
                    make_copy_cell(ds),           # col 3: DUMMY + COURSE
                    make_copy_cell(ds),           # col 4: DUMMY + COURSE
                ]

                tbl = Table(
                    [row_cells],
                    colWidths=[cell_w] * 4,
                    rowHeights=[cell_h / ROWS_PER_PAGE * (STUDS_PER_PAGE / len(page_studs) if len(page_studs) < STUDS_PER_PAGE else 1)],
                )
                tbl.setStyle(TableStyle([
                    ('ALIGN',     (0,0), (-1,-1), 'CENTER'),
                    ('VALIGN',    (0,0), (-1,-1), 'MIDDLE'),
                    ('GRID',      (0,0), (-1,-1), 0.8, colors.HexColor('#cccccc')),
                    ('LINEAFTER', (0,0), (0,-1),  1.5, NAVY),
                    ('TOPPADDING',(0,0), (-1,-1), 6),
                    ('BOTTOMPADDING',(0,0),(-1,-1), 6),
                ]))
                story.append(tbl)

        if page_idx < total_pages - 1:
            story.append(PageBreak())

    doc.build(story)
    buf.seek(0)
    filename = f'ESE_Stickers_{code}_{datetime.date.today()}.pdf'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/pdf')


# ── Student Results API ───────────────────────────────────────────────────────

@api.route('/results/my', methods=['GET'])
@login_required
def get_my_results():
    """
    Fetch results for the currently logged-in student.
    Links Student -> DummySticker -> FoilMark -> Course.
    """
    if current_user.role != 'student':
        # For testing/admin, allow passing regno
        regno = request.args.get('regno')
        if not regno:
            return jsonify({'message': 'Access denied'}), 403
        student = Student.query.filter_by(register_number=regno).first()
    else:
        student = Student.query.filter_by(register_number=current_user.username).first()

    if not student:
        return jsonify({'message': 'Student record not found'}), 404

    # Join results
    # 1. Get all dummy stickers for this student
    stickers = DummySticker.query.filter_by(student_id=student.id).all()
    
    results = []
    for s in stickers:
        # 2. Get marks for each dummy/course
        mark = FoilMark.query.filter_by(
            dummy_number=s.dummy_number,
            course_id=s.exam_schedule.course_id
        ).first()

        if mark:
            results.append({
                'semester':     s.exam_schedule.course.semester,
                'course_code':  s.exam_schedule.course.course_code,
                'course_title': s.exam_schedule.course.course_title,
                'grade':        mark.grade or '-',
            })

    # Sort by semester then code
    results.sort(key=lambda x: (x['semester'], x['course_code']))

    ay = AcademicYear.query.filter_by(is_current=True).first()
    session_label = f"{ay.semester if ay else 'NOV/DEC'} {ay.label.split('-')[0] if ay else '2024'}"

    ay = AcademicYear.query.filter_by(is_current=True).first()
    session_label = f"{ay.semester if ay else 'NOV/DEC'} {ay.label.split('-')[0] if ay else '2024'}"

    return jsonify({
        'student_name':  student.name,
        'register_no':   student.register_number,
        'session_title': f"END SEMESTER EXAMINATION RESULTS - {session_label}",
        'results':       results
    })


# ── Hall Ticket PDF Generation ────────────────────────────────────────────────

@api.route('/ese/hallticket-pdf', methods=['GET'])
@login_required
def get_hallticket_pdf():
    ay = AcademicYear.query.filter_by(is_current=True).first()
    if not ay or not ay.hall_ticket_published:
        return jsonify({
            'message': 'Hall Tickets are not yet released for the current session.',
            'status': 'coming_soon'
        }), 403

    regno = request.args.get('regno')
    if current_user.role == 'student':
        regno = current_user.username
    
    if not regno:
        return jsonify({'message': 'Register number required'}), 400
    
    student = Student.query.filter_by(register_number=regno).first()
    if not student:
        return jsonify({'message': 'Student not found'}), 404

    ay = AcademicYear.query.filter_by(is_current=True).first()
    session_label = f"{ay.semester if ay else 'NOV/DEC'} {ay.label.split('-')[0] if ay else '2025'}"

    # Fetch Exams for this student
    schedules = ExamSchedule.query.join(Course).filter(
        Course.department == student.department,
        Course.semester == student.semester,
        ExamSchedule.academic_year_id == (ay.id if ay else None)
    ).all()

    if not schedules:
        return jsonify({'message': 'No exam schedules found for this student in the current session.'}), 404

    # Sort schedules by date
    schedules.sort(key=lambda x: x.exam_date if x.exam_date else datetime.date.max)

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    
    # Custom Page Border using Canvas
    def add_page_border(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColor(colors.black)
        canvas.setLineWidth(2)
        # Main border
        canvas.rect(20, 20, A4[0]-40, A4[1]-40)
        # Inner border (double effect)
        canvas.setLineWidth(0.5)
        canvas.rect(25, 25, A4[0]-50, A4[1]-50)
        canvas.restoreState()

    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    
    story = []

    # 1. Header (Logo | College Info | Quality Logo)
    logo_path = os.path.join(os.getcwd(), 'frontend', 'public', 'logo.png')
    
    # Placeholder for the right side quality logo circle
    header_data = [
        [
            RLImage(logo_path, width=0.8*inch, height=0.8*inch) if os.path.exists(logo_path) else "",
            Paragraph("<b>KINGS ENGINEERING COLLEGE</b><br/>"
                       "<font size=8>AN AUTONOMOUS INSTITUTION<br/>"
                       "ACCREDITED WITH NAAC AND AFFILIATED TO ANNA UNIVERSITY<br/>"
                       "Chennai-Bangalore Highway, Irungattukottai, Chennai - 602 117.<br/>"
                       "Ph.: 044 - 71224401 -08. Fax:044 - 71224410</font>", 
                       ParagraphStyle('h_center', parent=styles['Normal'], alignment=1, leading=11)),
            Paragraph("<font size=6>INTERNAL QUALITY ASSURANCE CELL</font>", 
                      ParagraphStyle('h_right', parent=styles['Normal'], alignment=1, fontSize=6, borderPadding=10))
        ]
    ]
    header_table = Table(header_data, colWidths=[1.1*inch, 4.3*inch, 1.1*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('LINEBELOW', (0,0), (-1,-1), 1, colors.black),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 15))

    # 2. Titles
    story.append(Paragraph(f"<b>END SEMESTER EXAMINATIONS – {session_label}</b>", 
                          ParagraphStyle('t1', parent=styles['Normal'], alignment=1, fontSize=11)))
    story.append(Spacer(1, 8))
    story.append(Paragraph("<b>HALL TICKET</b>", 
                          ParagraphStyle('t2', parent=styles['Normal'], alignment=1, fontSize=15)))
    story.append(Spacer(1, 15))

    # 3. Student Info Table
    info_data = [
        ["Register Number", student.register_number, "Current Semester", str(student.semester), ""],
        ["Name of the Student", student.name, "Date of Birth", student.dob or "09/08/2004", "P h o t o"],
        ["Degree & Branch", f"{student.degree or 'B.E.'} {student.department}", "Regulation", student.regulation or "2021", ""]
    ]
    info_table = Table(info_data, colWidths=[1.4*inch, 2.6*inch, 1.2*inch, 1.2*inch, 1*inch])
    info_table.setStyle(TableStyle([
        ('GRID', (0,0), (-2,-1), 1.2, colors.black),
        ('GRID', (4,0), (4,2), 1.2, colors.black), # Photo box
        ('SPAN', (4,0), (4,2)),
        ('ALIGN', (4,0), (4,2), 'CENTER'),
        ('VALIGN', (4,0), (4,2), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 20))

    # 4. Exam Schedule Table (2 columns)
    sched_header = [
        Paragraph("<b>Sem</b>", styles['Normal']), 
        Paragraph("<b>Course Code</b>", styles['Normal']), 
        Paragraph("<b>Course Title</b>", styles['Normal']),
        Paragraph("<b>Sem</b>", styles['Normal']), 
        Paragraph("<b>Course Code</b>", styles['Normal']), 
        Paragraph("<b>Course Title</b>", styles['Normal'])
    ]
    sched_rows = [sched_header]
    
    half = (len(schedules) + 1) // 2
    for i in range(half):
        row = []
        # Col 1
        s1 = schedules[i]
        dt_str = s1.exam_date.strftime('%d.%m.%Y %a') if s1.exam_date else ""
        row.extend([str(s1.course.semester), s1.course.course_code, Paragraph(f"<font size=8>{s1.course.course_title} <font color=grey>[{dt_str} - {s1.session}]</font></font>", styles['Normal'])])
        
        # Col 2
        if i + half < len(schedules):
            s2 = schedules[i + half]
            dt_str2 = s2.exam_date.strftime('%d.%m.%Y %a') if s2.exam_date else ""
            row.extend([str(s2.course.semester), s2.course.course_code, Paragraph(f"<font size=8>{s2.course.course_title} <font color=grey>[{dt_str2} - {s2.session}]</font></font>", styles['Normal'])])
        else:
            row.extend(["", "", ""])
        sched_rows.append(row)

    while len(sched_rows) < 18:
        sched_rows.append(["", "", "", "", "", ""])

    sched_table = Table(sched_rows, colWidths=[0.4*inch, 0.9*inch, 2.3*inch, 0.4*inch, 0.9*inch, 2.3*inch])
    sched_table.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.8, colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (1,-1), 'CENTER'),
        ('ALIGN', (3,0), (4,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(sched_table)
    story.append(Spacer(1, 10))

    # 5. Footer Summary
    story.append(Paragraph(f"<b>No of Courses Registered: {len(schedules)}</b>", styles['Normal']))
    story.append(Spacer(1, 10))
    
    note_box_data = [[
        Paragraph("<b>NOTE :</b><br/>"
                  "1. In case of candidates who have been Readmitted/Transferred, this Hall Ticket is valid only if the candidate's admission is approved by the Commissioner of Technical Education, Government of Tamil Nadu, Chennai and the Registrar, Anna University, Chennai. If any candidate appears for the examination without the approval, the examination written by the candidate will be invalidated.<br/>"
                  "2. Correction in the Name / Date of Birth and missing of Photograph or incorrect Photograph, if any is to be updated in the Web Portal when it is opened for correction. The date and time for the correction will be informed.<br/>"
                  "3. Instructions printed overleaf are to be followed strictly.", 
                  ParagraphStyle('note_box', parent=styles['Normal'], fontSize=7.5, leading=10))
    ]]
    note_table = Table(note_box_data, colWidths=[7.2*inch])
    note_table.setStyle(TableStyle([
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(note_table)
    
    story.append(Spacer(1, 5))
    story.append(Paragraph(f"<b>Generated On: {datetime.date.today().strftime('%d.%m.%Y')}</b>", ParagraphStyle('gen', parent=styles['Normal'], fontSize=9)))
    story.append(Spacer(1, 20))

    sig_data = [
        ["Signature of the Candidate", "Signature of the Principal", "Controller of Examinations with Seal"]
    ]
    sig_table = Table(sig_data, colWidths=[2.4*inch, 2.4*inch, 2.4*inch])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8.5),
        ('TOPPADDING', (0,0), (-1,-1), 35),
        ('LINEABOVE', (0,0), (-1,0), 0.5, colors.grey),
    ]))
    story.append(sig_table)

    # 6. Instructions Page (Page 2)
    story.append(PageBreak())
    story.append(Paragraph("<b>INSTRUCTIONS TO THE CANDIDATE</b>", ParagraphStyle('inst_title', parent=styles['Normal'], alignment=1, fontSize=12)))
    story.append(Spacer(1, 15))
    
    inst_style = ParagraphStyle('inst', parent=styles['Normal'], fontSize=9, leading=13, alignment=4) # justified
    insts = [
        "1. Admission to the Examination is provisional.",
        "2. The Hall Ticket is issued subject to the candidate satisfying the Attendance and other requirements as per Rules, Regulations and Instructions by the Institution from time to time. If later, it is found that the candidate fails to comply with the above requirements, the examinations written by the candidate will be treated as cancelled.",
        "3. A seat marked with Register number will be provided to each candidate. Candidate will occupy the allotted seat at least 10 minutes before the commencement of the examination. In no case, candidate shall be allowed to occupy a seat other than the seat allotted to him.",
        "4. Normally the Candidate will not be permitted to enter the hall after the commencement of the examination. Only on extraordinary circumstances, the candidates will be permitted during the first thirty minutes of the examination after obtaining the written permission from the Principal / CS. Under any circumstances the Candidate shall not be permitted to enter the hall after the expiry of first thirty minutes.",
        "5. Candidate shall not be allowed to leave the examination hall before the expiry of 45 minutes from the commencement of examination. The candidate who leaves the examination hall during the period allotted for a paper will not be allowed to re-enter the hall within that period.",
        "6. Candidate who is suffering from infectious diseases of any kind shall not be admitted to the examination hall.",
        "7. Candidate is strictly prohibited from smoking inside the examination hall.",
        "8. Strict silence should be maintained in the examination hall.",
        "9. Candidate is required to bring his/her own pens, pencils and erasers. Candidate should use only blue or black ink while answering his/her papers.",
        "10. Before proceeding to answer the paper, the candidate should write his/her register number, semester, subject and date of the examination at the appropriate space provided in the first page of the answer book and nowhere else in the answer book or in any additional attachment like drawing sheet, smith chart etc.",
        "11. If a candidate writes his/her register number on any part of the answer book/sheets other than the one provided for or puts any special mark or writes anything which may disclose, in any way, the identity of the Candidate/College, he/she will render himself/herself liable for disciplinary action.",
        "12. Writing of wrong register number in the answer book will entail rejection of the answer book.",
        "13. Candidate is not allowed to exceed the prescribed time assigned to each paper.",
        "14. Candidate shall not talk/ask questions of any kind during the examination.",
        "15. Candidate shall not carry any written / printed matter, any paper material, cell phone, pen drive, ipad, programmable calculator, any unauthorized data sheet / table into the examination hall and if anything is found in his/her possession his/her shall be liable for disciplinary action.",
        "16. No Candidate shall pass any part or whole of answer papers or question papers to any other candidate. No candidate shall allow another candidate to copy from his/her answer paper or copy from the answer paper of another candidate. If found committing such malpractice, the involved candidates shall be liable for disciplinary action.",
        "17. Candidate found guilty of using unfair means of any nature shall be liable for disciplinary action.",
        "18. Candidate will have to hand over the answer book to the Invigilator / Chief Superintendent before leaving the examination hall.",
        "19. Candidate should produce the hall ticket on demand by the Invigilator / Chief Superintendent / Anna University Representative / Squad members.",
        "20. Candidate shall not write anything in the Hall Ticket.",
        "21. Candidate shall write only the Register No. in the space provided in the Question Paper. Any other writings in the Question Paper is prohibited and punishable."
    ]
    for i in insts:
        story.append(Paragraph(i, inst_style))
        story.append(Spacer(1, 5))

    story.append(Spacer(1, 20))
    story.append(Paragraph("<b>CONTROLLER OF EXAMINATIONS</b>", ParagraphStyle('coe', parent=styles['Normal'], alignment=2, fontSize=10)))

    # Build with border
    doc.build(story, onFirstPage=add_page_border, onLaterPages=add_page_border)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f'HallTicket_{regno}.pdf', mimetype='application/pdf')


# ── Publication Controls ──────────────────────────────────────────────────────

@api.route('/ese/toggle-publish', methods=['POST'])
@login_required
def toggle_publish():
    if current_user.role != 'admin':
        logging.warning(f"Unauthorized publish attempt by {current_user.username}")
        return jsonify({'message': 'Access denied'}), 403
    
    data = request.json
    target = data.get('target') # 'hall_ticket' or 'results'
    value = data.get('value')   # True or False
    
    ay = AcademicYear.query.filter_by(is_current=True).first()
    if not ay:
        return jsonify({'message': 'Current academic year not found'}), 404
    
    try:
        if target == 'hall_ticket':
            ay.hall_ticket_published = value
        elif target == 'results':
            ay.results_published = value
        else:
            return jsonify({'message': 'Invalid target'}), 400
            
        db.session.commit()
        logging.info(f"ADMIN {current_user.username} updated {target} visibility to {value}")
        return jsonify({
            'message': f"{target.replace('_', ' ').title()} visibility updated.",
            'hall_ticket_published': ay.hall_ticket_published,
            'results_published': ay.results_published
        })
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating publication status: {str(e)}")
        return jsonify({'message': 'Failed to update status due to server error'}), 500

@api.route('/ese/publication-status', methods=['GET'])
@login_required
def get_publication_status():
    ay = AcademicYear.query.filter_by(is_current=True).first()
    return jsonify({
        'hall_ticket_published': ay.hall_ticket_published if ay else False,
        'results_published': ay.results_published if ay else False
    })


